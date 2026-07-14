from decimal import Decimal

from openpyxl import Workbook, load_workbook

from bdap_app.core.automation import fill_accounting_analysis
from bdap_app.core.automation_fcde import populate_fcde_sheet_in_workbook
from bdap_app.support.default_mappings import TEMPLATE_SOURCES


def _workbook_with_sheets(count: int) -> Workbook:
    wb = Workbook()
    wb.active.title = "Sheet1"
    for idx in range(2, count + 1):
        wb.create_sheet(f"Sheet{idx}")
    return wb


def _analysis_workbook_with_fcde_year(year: int) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "FCDE"
    ws["D1"] = year
    return wb


def _save_rendiconto(path, title_one_row: int = 19, title_one_value: int = 100) -> None:
    wb = _workbook_with_sheets(16)
    ws_importo = wb.worksheets[10]
    ws_importo["B27"] = 30

    ws_residui = wb.worksheets[15]
    ws_residui["B19"] = "Totale Titolo 1: Entrate correnti di natura tributaria, contributiva e perequativa"
    ws_residui["D19"] = 100
    ws_residui["B35"] = "Totale Titolo 3: Entrate extratributarie"
    ws_residui["D35"] = 200

    if title_one_row != 19:
        ws_residui["B19"] = "Voce diversa"
        ws_residui["D19"] = 999
        ws_residui.cell(row=title_one_row, column=2).value = (
            "Totale Titolo 1: Entrate correnti di natura tributaria, contributiva e perequativa"
        )
        ws_residui.cell(row=title_one_row, column=4).value = title_one_value

    wb.save(path)


def test_populate_fcde_sheet_uses_rendiconto_residui_mappings(tmp_path) -> None:
    analysis_wb = _analysis_workbook_with_fcde_year(2024)
    bdap_path = tmp_path / "Rendiconto 2024.xlsx"
    _save_rendiconto(bdap_path)

    percentage, cell_ref = populate_fcde_sheet_in_workbook(
        analysis_wb,
        bdap_path,
        2024,
    )

    ws_fcde = analysis_wb["FCDE"]
    assert ws_fcde["D2"].value == Decimal("30")
    assert ws_fcde["D3"].value == Decimal("100")
    assert ws_fcde["D4"].value == Decimal("200")
    assert ws_fcde["D5"].value == Decimal("300")
    assert ws_fcde["D6"].value == Decimal("10")
    assert ws_fcde["D7"].value == "ordinario"
    assert percentage == Decimal("10")
    assert cell_ref == "FCDE!D6"


def test_populate_fcde_sheet_adjusts_rendiconto_residui_row_by_label(tmp_path) -> None:
    analysis_wb = _analysis_workbook_with_fcde_year(2024)
    bdap_path = tmp_path / "Rendiconto 2024.xlsx"
    _save_rendiconto(bdap_path, title_one_row=20, title_one_value=123)

    populate_fcde_sheet_in_workbook(
        analysis_wb,
        bdap_path,
        2024,
    )

    ws_fcde = analysis_wb["FCDE"]
    assert ws_fcde["D3"].value == Decimal("123")
    assert ws_fcde["D5"].value == Decimal("323")


def test_populate_fcde_sheet_does_not_fallback_to_questionario_residui(tmp_path) -> None:
    analysis_wb = _analysis_workbook_with_fcde_year(2024)
    bdap_path = tmp_path / "Rendiconto 2024.xlsx"
    wb_bdap = _workbook_with_sheets(16)
    wb_bdap.worksheets[10]["B27"] = 30
    wb_bdap.save(bdap_path)

    wb_questionario = Workbook()
    ws_questionario = wb_questionario.active
    ws_questionario.title = "SEZ. II -DATI RISULT. FINAN"
    ws_questionario["D69"] = 2024
    ws_questionario["B72"] = "TITOLO I"
    ws_questionario["D72"] = 999
    ws_questionario["B74"] = "TITOLO III"
    ws_questionario["D74"] = 888
    wb_questionario.save(tmp_path / "Questionario 2024.xlsx")

    percentage, cell_ref = populate_fcde_sheet_in_workbook(
        analysis_wb,
        bdap_path,
        2024,
    )

    ws_fcde = analysis_wb["FCDE"]
    assert ws_fcde["D2"].value == Decimal("30")
    assert ws_fcde["D3"].value is None
    assert ws_fcde["D4"].value is None
    assert ws_fcde["D5"].value is None
    assert ws_fcde["D6"].value is None
    assert percentage is None
    assert cell_ref == "FCDE!D6"


def test_fill_accounting_analysis_writes_fcde_ratio_from_fcde_sheet(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "Analisi Contabile"
    ws_analysis["C1"] = 2023
    ws_analysis["D1"] = 2024
    ws_analysis["C13"] = "FCDE IN RAPPORTO A RESIDUI ATTIVI %\n\n(FCDE/Totale residui attivi >= 30%)"

    ws_fcde = wb_analysis.create_sheet("FCDE")
    ws_fcde["D1"] = 2024
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rendiconto 2024.xlsx"
    _save_rendiconto(bdap_path)

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2024,
        analysis_sheet="Analisi Contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["Analisi Contabile"]
        ws_fcde_result = wb_result["FCDE"]
        assert updated == 1
        assert unresolved == 0
        assert ws_fcde_result["D6"].value == 10
        assert ws_result["D13"].value == "10,00"
        assert ws_result["D13"].comment is not None
        assert "foglio FCDE, cella FCDE!D6" in ws_result["D13"].comment.text
    finally:
        wb_result.close()


def test_default_mappings_keep_only_dicuifcde_for_fcde_amount() -> None:
    assert "dicuifcde" in TEMPLATE_SOURCES
    assert "importofcde" not in TEMPLATE_SOURCES
