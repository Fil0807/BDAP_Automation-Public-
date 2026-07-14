from openpyxl import Workbook, load_workbook

from bdap_app.support.controlli_post import (
    REPORT_SHEET_NAME,
    collect_controlli_post_report,
    write_controlli_post_sheet,
)


def _cell_fill_rgb(cell) -> str:
    return str(cell.fill.fgColor.rgb or "")


def _save_controlli_post(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Controlli Post"
    ws.append(["ID", "AREA TEMATICA", "CONTROLLO ", "ESITO"])
    ws.append([
        1,
        "Equilibri di bilancio",
        "Domanda con esito nel testo\nLa risposta alla domanda e NO\nCriticità contabile",
        None,
    ])
    ws.append([
        2,
        "Organismi partecipati",
        "Domanda con esito in colonna",
        "Warning",
    ])
    ws.append([
        3,
        "Equilibri di bilancio",
        "Il revisore ha segnalato criticità/anomalie?\nLa risposta alla domanda e NO",
        None,
    ])
    ws.append(["Warning", "Conteggio Warning", "", 1])
    wb.save(path)


def test_collect_controlli_post_report_reads_reported_issues(tmp_path) -> None:
    _save_controlli_post(tmp_path / "Controlli Post - Questionario 2024.xlsx")

    report = collect_controlli_post_report(tmp_path, years=[2024])

    assert len(report.files) == 1
    assert len(report.issues) == 2
    assert [issue.severity for issue in report.issues] == ["Criticità contabile", "Warning"]
    assert report.issues[0].year == 2024
    assert "Criticità contabile" not in report.issues[0].description
    assert report.issues[1].area == "Organismi partecipati"


def test_write_controlli_post_sheet_fills_existing_template_sheet(tmp_path) -> None:
    controlli_path = tmp_path / "Controlli Post - Questionario 2024.xlsx"
    _save_controlli_post(controlli_path)
    report = collect_controlli_post_report(tmp_path, years=[2024])

    output_path = tmp_path / "Template_Analisi_compilato.xlsx"
    wb = Workbook()
    wb.active.title = "ANALISI CONTABILE"
    ws_report = wb.create_sheet(REPORT_SHEET_NAME)
    ws_report["A1"] = "Criticità Controlli Post"
    ws_report["A2"] = "File analizzati: "
    ws_report["A3"] = "Segnalazioni"
    ws_report.append(["Anno", "Livello", "Area tematica", "ID", "Controllo", "File sorgente", "Riga"])
    wb.save(output_path)

    assert write_controlli_post_sheet(output_path, report) is True

    result = load_workbook(output_path)
    try:
        assert REPORT_SHEET_NAME in result.sheetnames
        ws = result[REPORT_SHEET_NAME]
        assert ws["A1"].value == "Criticità Controlli Post"
        assert ws["A3"].value == "Segnalazioni: 2; Criticità contabili: 1; Warning: 1"
        assert ws["A5"].value == 2024
        assert ws["B5"].value == "Criticità contabile"
        assert ws["B6"].value == "Warning"
        assert _cell_fill_rgb(ws["B5"]).endswith("F4CCCC")
        assert _cell_fill_rgb(ws["B6"]).endswith("FFF2CC")
        assert _cell_fill_rgb(ws["E5"]).endswith("F4CCCC")
        assert _cell_fill_rgb(ws["E6"]).endswith("FFF2CC")
    finally:
        result.close()


def test_write_controlli_post_sheet_does_not_create_missing_template_sheet(tmp_path) -> None:
    controlli_path = tmp_path / "Controlli Post - Questionario 2024.xlsx"
    _save_controlli_post(controlli_path)
    report = collect_controlli_post_report(tmp_path, years=[2024])

    output_path = tmp_path / "Template_Analisi_compilato.xlsx"
    wb = Workbook()
    wb.active.title = "ANALISI CONTABILE"
    wb.save(output_path)

    assert write_controlli_post_sheet(output_path, report) is False

    result = load_workbook(output_path)
    try:
        assert REPORT_SHEET_NAME not in result.sheetnames
    finally:
        result.close()
