from openpyxl import Workbook, load_workbook

from bdap_app.core.automation import fill_accounting_analysis
from bdap_app.support import default_mappings


def _workbook_with_sheets(count: int) -> Workbook:
    wb = Workbook()
    wb.active.title = "Sheet1"
    for idx in range(2, count + 1):
        wb.create_sheet(f"Sheet{idx}")
    return wb


def test_relazione_source_leaves_cell_empty_and_counts_unresolved(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2024
    ws_analysis["A2"] = "spesadelpersonalesostenuta"
    ws_analysis["B2"] = "valore precedente"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rend. 2024.xlsx"
    Workbook().save(bdap_path)

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2024,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 0
        assert unresolved == 1
        assert ws_result["B2"].value is None
        assert ws_result["B2"].comment is None
    finally:
        wb_result.close()


def test_unmapped_row_is_not_filled_by_global_label_search(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2024
    ws_analysis["A2"] = "Voce non mappata"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rend. 2024.xlsx"
    wb_bdap = Workbook()
    ws_bdap = wb_bdap.active
    ws_bdap["A1"] = "Voce non mappata"
    ws_bdap["B1"] = 123
    wb_bdap.save(bdap_path)

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2024,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 0
        assert unresolved == 0
        assert ws_result["B2"].value is None
        assert ws_result["B2"].comment is None
    finally:
        wb_result.close()


def test_explicit_bdap_reference_is_still_compiled(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2024
    ws_analysis["A2"] = "Voce con riferimento"
    ws_analysis["B2"] = "BDAP foglio 1, B1"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rend. 2024.xlsx"
    wb_bdap = Workbook()
    ws_bdap = wb_bdap.active
    ws_bdap["B1"] = 456
    wb_bdap.save(bdap_path)

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2024,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 1
        assert unresolved == 0
        assert ws_result["B2"].value == "456,00"
        assert ws_result["B2"].comment is not None
    finally:
        wb_result.close()


def test_fill_accounting_analysis_reports_row_progress(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2024
    ws_analysis["A2"] = "Voce con riferimento"
    ws_analysis["B2"] = "BDAP foglio 1, B1"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rend. 2024.xlsx"
    wb_bdap = Workbook()
    ws_bdap = wb_bdap.active
    ws_bdap["B1"] = 456
    wb_bdap.save(bdap_path)

    progress_events = []
    fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2024,
        analysis_sheet="analisi contabile",
        progress=lambda current, total: progress_events.append((current, total)),
    )

    assert progress_events[0] == (0, 2)
    assert progress_events[-1] == (2, 2)
    assert any(current > 0 for current, _total in progress_events)


def test_mapped_cell_ref_still_works_when_expected_label_does_not_match(tmp_path) -> None:
    original_sources = dict(default_mappings.TEMPLATE_SOURCES)
    default_mappings.TEMPLATE_SOURCES["vocecustom"] = {
        "cell_ref": "B2",
        "expected_label": "etichetta non presente",
        "optional": True,
        "sheet_idx": "1",
        "source_workbook": "rendiconto",
    }

    try:
        analysis_path = tmp_path / "analysis.xlsx"
        wb_analysis = Workbook()
        ws_analysis = wb_analysis.active
        ws_analysis.title = "analisi contabile"
        ws_analysis["A1"] = "Voce"
        ws_analysis["B1"] = 2024
        ws_analysis["A2"] = "Voce custom"
        wb_analysis.save(analysis_path)

        bdap_path = tmp_path / "Rend. 2024.xlsx"
        wb_bdap = Workbook()
        ws_bdap = wb_bdap.active
        ws_bdap["A2"] = "altra etichetta"
        ws_bdap["B2"] = 789
        wb_bdap.save(bdap_path)

        updated, unresolved = fill_accounting_analysis(
            analysis_path=analysis_path,
            bdap_path=bdap_path,
            output_path=analysis_path,
            year=2024,
            analysis_sheet="analisi contabile",
        )

        wb_result = load_workbook(analysis_path)
        try:
            ws_result = wb_result["analisi contabile"]
            assert updated == 1
            assert unresolved == 0
            assert ws_result["B2"].value == "789,00"
        finally:
            wb_result.close()
    finally:
        default_mappings.TEMPLATE_SOURCES.clear()
        default_mappings.TEMPLATE_SOURCES.update(original_sources)


def test_fondo_contenzioso_uses_list_expected_label_to_adjust_row(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2024
    ws_analysis["A2"] = "Fondo contenzioso"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rend. 2024.xlsx"
    wb_bdap = _workbook_with_sheets(12)
    ws_source = wb_bdap.worksheets[11]
    ws_source["E11"] = 2024
    ws_source["I16"] = 111
    ws_source["E18"] = 999
    ws_source["B18"] = "Totale Fondo contenzioso"
    ws_source["I18"] = 222
    wb_bdap.save(bdap_path)

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2024,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 1
        assert unresolved == 0
        assert ws_result["B2"].value == "222,00"
    finally:
        wb_result.close()


def test_potenziali_passivita_uses_list_expected_label_to_adjust_row(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2024
    ws_analysis["A2"] = "Potenziali passivita"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rend. 2024.xlsx"
    wb_bdap = _workbook_with_sheets(12)
    ws_source = wb_bdap.worksheets[11]
    ws_source["E11"] = 2024
    ws_source["I26"] = 333
    ws_source["E28"] = 999
    ws_source["D28"] = "FONDO ACCANTONAMENTO PASSIVITA POTENZIALI"
    ws_source["I28"] = 444
    wb_bdap.save(bdap_path)

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2024,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 1
        assert unresolved == 0
        assert ws_result["B2"].value == "444,00"
    finally:
        wb_result.close()


def test_potenziali_passivita_defaults_to_zero_when_label_is_missing(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2022
    ws_analysis["A2"] = "Potenziali passivita"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rend. 2022.xlsx"
    wb_bdap = _workbook_with_sheets(12)
    ws_source = wb_bdap.worksheets[11]
    ws_source["I26"] = 12345
    ws_source["B28"] = "Totale risorse accantonate"
    ws_source["I28"] = 67890
    wb_bdap.save(bdap_path)

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2022,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 1
        assert unresolved == 0
        assert ws_result["B2"].value == "0,00"
    finally:
        wb_result.close()


def test_residui_datati_prefer_exact_year_questionario_table(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2022
    ws_analysis["A2"] = "RESIDUI ATTIVI DATATI precedenti 2021 (TITOLO I)"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rend. 2022.xlsx"
    Workbook().save(bdap_path)

    old_questionario = tmp_path / "Questionario 2022.xlsx"
    wb_old = Workbook()
    ws_old = wb_old.active
    ws_old.title = "SEZ. II -DATI RISULT. FINAN"
    ws_old["D69"] = 2021
    ws_old["E69"] = 2022
    ws_old["B72"] = "Titolo I"
    ws_old["E72"] = 111
    wb_old.save(old_questionario)

    latest_questionario = tmp_path / "Questionario 2024.xlsx"
    wb_latest = Workbook()
    ws_latest = wb_latest.active
    ws_latest.title = "SEZ. II -DATI RISULT. FINAN"
    ws_latest["D69"] = 2021
    ws_latest["E69"] = 2022
    ws_latest["F69"] = 2023
    ws_latest["G69"] = 2024
    ws_latest["B72"] = "Titolo I"
    ws_latest["E72"] = 222
    wb_latest.save(latest_questionario)

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2022,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 1
        assert unresolved == 0
        assert ws_result["B2"].value == "TITOLO I: 111,00"
    finally:
        wb_result.close()


def test_residui_datati_resolve_offset_formula_when_cache_is_empty(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2023
    ws_analysis["A2"] = "RESIDUI ATTIVI DATATI precedenti 2021 (TITOLO I)"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rend. 2023.xlsx"
    Workbook().save(bdap_path)

    questionario_path = tmp_path / "Questionario 2023.xlsx"
    wb_questionario = Workbook()
    ws_questionario = wb_questionario.active
    ws_questionario.title = "SEZ. II-DATI RISULT. FINANZ."
    ws_data = wb_questionario.create_sheet("DATI")
    ws_questionario["C71"] = "Esercizi"
    ws_questionario["D71"] = 2021
    ws_questionario["E71"] = 2022
    ws_questionario["F71"] = 2023
    ws_questionario["B73"] = "Titolo I"
    ws_questionario["F73"] = '=IF(ISBLANK(OFFSET(DATI!$A$2,0,5)),"",OFFSET(DATI!$A$2,0,5))'
    ws_data["F2"] = 333
    wb_questionario.save(questionario_path)

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2023,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 1
        assert unresolved == 0
        assert ws_result["B2"].value == "TITOLO I: 333,00"
    finally:
        wb_result.close()


def test_residui_datati_use_title_row_and_year_column(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2024
    ws_analysis["A2"] = "RESIDUI ATTIVI DATATI precedenti 2021 (TITOLO I)"
    ws_analysis["A3"] = "RESIDUI ATTIVI DATATI precedenti 2021 (TITOLO III)"
    ws_analysis["A4"] = "RESIDUI ATTIVI DATATI precedenti 2021 (TITOLO IV)"
    ws_analysis["A5"] = "RESIDUI PASSIVI DATATI"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rend. 2024.xlsx"
    Workbook().save(bdap_path)

    questionario_path = tmp_path / "0000_Questionario 2024.xlsx"
    wb_questionario = Workbook()
    ws_questionario = wb_questionario.active
    ws_questionario.title = "SEZ. II -DATI RISULT. FINAN"

    ws_questionario["C69"] = "precedenti 2021"
    ws_questionario["D69"] = 2021
    ws_questionario["E69"] = 2022
    ws_questionario["F69"] = 2023
    ws_questionario["G69"] = 2024
    ws_questionario["C71"] = 101
    ws_questionario["B72"] = "Titolo I"
    ws_questionario["G72"] = 901
    ws_questionario["C73"] = 303
    ws_questionario["B74"] = "Titolo III"
    ws_questionario["G74"] = 903
    ws_questionario["C74"] = 404
    ws_questionario["B75"] = "Titolo IV"
    ws_questionario["G75"] = 904

    ws_questionario["C81"] = "precedenti 2021"
    ws_questionario["D81"] = 2021
    ws_questionario["E81"] = 2022
    ws_questionario["F81"] = 2023
    ws_questionario["G81"] = 2024
    ws_questionario["C83"] = 505
    ws_questionario["B86"] = "Titolo I"
    ws_questionario["G86"] = 905
    wb_questionario.save(questionario_path)

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2024,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 4
        assert unresolved == 0
        assert ws_result["B2"].value == "TITOLO I: 901,00"
        assert ws_result["B3"].value == "TITOLO III: 903,00"
        assert ws_result["B4"].value == "TITOLO IV: 904,00"
        assert ws_result["B5"].value == "TITOLO I: 905,00"
    finally:
        wb_result.close()


def test_residui_attivi_fallback_to_questionario_dati_cassa(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2021
    ws_analysis["A2"] = "RESIDUI ATTIVI DATATI precedenti 2021 (TITOLO I)"
    ws_analysis["A3"] = "RESIDUI ATTIVI DATATI precedenti 2021 (TITOLO III)"
    ws_analysis["A4"] = "RESIDUI ATTIVI DATATI precedenti 2021 (TITOLO IV)"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rend. 2021.xlsx"
    Workbook().save(bdap_path)

    questionario_path = tmp_path / "Questionario 2021.xlsx"
    wb_questionario = Workbook()
    ws_questionario = wb_questionario.active
    ws_questionario.title = "SEZ. II - Dati Cassa"
    ws_questionario["B6"] = "Entrate Titolo 1.00"
    ws_questionario["F6"] = 110242.70
    ws_questionario["B10"] = "Entrate Titolo 3.00"
    ws_questionario["F10"] = 80419.38
    ws_questionario["B12"] = (
        "Entrate Titolo 4.02.06 - Contributi agli investimenti direttamente destinati "
        "al rimborso dei prestiti"
    )
    ws_questionario["F12"] = 3456.78
    wb_questionario.save(questionario_path)

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2021,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 3
        assert unresolved == 0
        assert ws_result["B2"].value == "TITOLO I: 110.242,70"
        assert ws_result["B3"].value == "TITOLO III: 80.419,38"
        assert ws_result["B4"].value == "TITOLO IV: 3.456,78"
    finally:
        wb_result.close()


def test_exact_year_questionario_wins_over_later_questionario_tables(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2021
    ws_analysis["A2"] = "RESIDUI ATTIVI DATATI precedenti 2021 (TITOLO I)"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rend. 2021.xlsx"
    Workbook().save(bdap_path)

    exact_questionario = tmp_path / "Questionario 2021.xlsx"
    wb_exact = Workbook()
    ws_exact = wb_exact.active
    ws_exact.title = "SEZ. II - Dati Cassa"
    ws_exact["B6"] = "Entrate Titolo 1.00"
    ws_exact["F6"] = 110242.70
    wb_exact.save(exact_questionario)

    later_questionario = tmp_path / "Questionario 2024.xlsx"
    wb_later = Workbook()
    ws_later = wb_later.active
    ws_later.title = "SEZ. II -DATI RISULT. FINAN"
    ws_later["D69"] = 2021
    ws_later["B72"] = "Titolo I"
    ws_later["D72"] = 999
    wb_later.save(later_questionario)

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2021,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 1
        assert unresolved == 0
        assert ws_result["B2"].value == "TITOLO I: 110.242,70"
        assert ws_result["B2"].comment is not None
        assert "Questionario 2021.xlsx" in ws_result["B2"].comment.text
    finally:
        wb_result.close()


def test_questionario_without_target_year_filename_is_not_used(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2021
    ws_analysis["A2"] = "RESIDUI ATTIVI DATATI precedenti 2021 (TITOLO I)"
    wb_analysis.save(analysis_path)

    bdap_dir = tmp_path / "case" / "datibdap"
    bdap_dir.mkdir(parents=True)
    bdap_path = bdap_dir / "Rend. 2021.xlsx"
    Workbook().save(bdap_path)

    later_questionario = bdap_dir / "Questionario 2024.xlsx"
    wb_later = Workbook()
    ws_later = wb_later.active
    ws_later.title = "SEZ. II -DATI RISULT. FINAN"
    ws_later["D69"] = 2021
    ws_later["B72"] = "Titolo I"
    ws_later["D72"] = 999
    wb_later.save(later_questionario)

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2021,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 0
        assert unresolved == 0
        assert ws_result["B2"].value is None
        assert ws_result["B2"].comment is None
    finally:
        wb_result.close()


def test_residui_passivi_fallback_to_questionario_dati_cassa_formula(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2021
    ws_analysis["A2"] = "RESIDUI PASSIVI DATATI"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rend. 2021.xlsx"
    Workbook().save(bdap_path)

    questionario_path = tmp_path / "Questionario 2021.xlsx"
    wb_questionario = Workbook()
    ws_questionario = wb_questionario.active
    ws_questionario.title = "SEZ. II - Dati Cassa"
    ws_data = wb_questionario.create_sheet("DATI")
    ws_questionario["B15"] = "Spese Titolo 1.00 - Spese correnti"
    ws_questionario["F15"] = '=IF(ISBLANK(OFFSET(DATI!$A$2,0,5)),"",OFFSET(DATI!$A$2,0,5))'
    ws_data["F2"] = 136770.95
    wb_questionario.save(questionario_path)

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2021,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 1
        assert unresolved == 0
        assert ws_result["B2"].value == "TITOLO I: 136.770,95"
    finally:
        wb_result.close()


def test_importo_cassa_vincolata_prefers_exact_year_questionario_table(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2022
    ws_analysis["A2"] = "Importo cassa vincolata"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rend. 2022.xlsx"
    Workbook().save(bdap_path)

    old_questionario = tmp_path / "Questionario 2022.xlsx"
    wb_old = Workbook()
    ws_old = wb_old.active
    ws_old.title = "SEZ. II-GEST. FIN. CASSA"
    ws_old["D6"] = 2022
    ws_old["B9"] = "di cui cassa vincolata"
    ws_old["D9"] = 57278
    wb_old.save(old_questionario)

    latest_questionario = tmp_path / "Questionario 2024.xlsx"
    wb_latest = Workbook()
    ws_latest = wb_latest.active
    ws_latest.title = "SEZ. II-GEST. FIN. CASSA"
    ws_latest["D6"] = 2022
    ws_latest["E6"] = 2023
    ws_latest["F6"] = 2024
    ws_latest["B9"] = "di cui cassa vincolata"
    ws_latest["D9"] = 999
    wb_latest.save(latest_questionario)

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2022,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 1
        assert unresolved == 0
        assert ws_result["B2"].value == "57.278,00"
    finally:
        wb_result.close()


def test_importo_cassa_vincolata_uses_dati_cassa_layout_2022(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2022
    ws_analysis["A2"] = "Importo cassa vincolata"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rend. 2022.xlsx"
    Workbook().save(bdap_path)

    questionario_path = tmp_path / "Questionario 2022.xlsx"
    wb_questionario = Workbook()
    ws_other = wb_questionario.active
    ws_other.title = "SEZ. II - GEST_FIN. 1_"
    ws_other["D10"] = 2022
    ws_other["B12"] = "altra cassa vincolata da ignorare"
    ws_other["D12"] = 999
    ws_questionario = wb_questionario.create_sheet(" SEZ. II -DATI CASSA_")
    ws_questionario["E62"] = 2020
    ws_questionario["F62"] = 2021
    ws_questionario["G62"] = 2022
    ws_questionario["C64"] = "di cui cassa vincolata"
    ws_questionario["G64"] = 57278
    wb_questionario.save(questionario_path)

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2022,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 1
        assert unresolved == 0
        assert ws_result["B2"].value == "57.278,00"
    finally:
        wb_result.close()


def test_importo_cassa_vincolata_ignores_controlli_post_questionario(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2022
    ws_analysis["A2"] = "Importo cassa vincolata"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rend. 2022.xlsx"
    Workbook().save(bdap_path)

    controlli_post_path = tmp_path / "Controlli Post - Questionario Bilancio Enti locali Consuntivo 2022.xlsx"
    wb_controlli = Workbook()
    wb_controlli.active.title = "Controlli Post"
    wb_controlli.save(controlli_post_path)

    questionario_path = tmp_path / "Questionario Bilancio Enti locali Consuntivo 2022.xlsx"
    wb_questionario = Workbook()
    ws_questionario = wb_questionario.active
    ws_questionario.title = " SEZ. II -DATI CASSA_"
    ws_questionario["E62"] = 2020
    ws_questionario["F62"] = 2021
    ws_questionario["G62"] = 2022
    ws_questionario["C64"] = "di cui cassa vincolata"
    ws_questionario["G64"] = 57278
    wb_questionario.save(questionario_path)

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2022,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 1
        assert unresolved == 0
        assert ws_result["B2"].value == "57.278,00"
        assert "Questionario Bilancio Enti locali Consuntivo 2022.xlsx" in ws_result["B2"].comment.text
    finally:
        wb_result.close()


def test_importo_cassa_vincolata_uses_gest_fin_layout_2023(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2023
    ws_analysis["A2"] = "Importo cassa vincolata"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rend. 2023.xlsx"
    Workbook().save(bdap_path)

    questionario_path = tmp_path / "Questionario 2023.xlsx"
    wb_questionario = Workbook()
    ws_questionario = wb_questionario.active
    ws_questionario.title = "SEZ. II - GEST_FIN. 1_"
    ws_questionario["D10"] = 2021
    ws_questionario["E10"] = 2022
    ws_questionario["F10"] = 2023
    ws_questionario["B12"] = "di cui cassa vincolata"
    ws_questionario["F12"] = 22398
    wb_questionario.save(questionario_path)

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2023,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 1
        assert unresolved == 0
        assert ws_result["B2"].value == "22.398,00"
    finally:
        wb_result.close()


def test_importo_cassa_vincolata_uses_dati_cassa_layout_2021(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2021
    ws_analysis["A2"] = "Importo cassa vincolata"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rend. 2021.xlsx"
    Workbook().save(bdap_path)

    questionario_path = tmp_path / "Questionario 2021.xlsx"
    wb_questionario = Workbook()
    ws_questionario = wb_questionario.active
    ws_questionario.title = "SEZ. II - Dati Cassa"
    ws_questionario["D61"] = 2019
    ws_questionario["E61"] = 2020
    ws_questionario["F61"] = 2021
    ws_questionario["B63"] = "di cui cassa vincolata"
    ws_questionario["F63"] = 70054.01
    wb_questionario.save(questionario_path)

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2021,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 1
        assert unresolved == 0
        assert ws_result["B2"].value == "70.054,01"
    finally:
        wb_result.close()


def test_importo_cassa_vincolata_uses_label_row_and_year_column(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2024
    ws_analysis["A2"] = "Importo cassa vincolata"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rend. 2024.xlsx"
    Workbook().save(bdap_path)

    questionario_path = tmp_path / "0000_Questionario 2024.xlsx"
    wb_questionario = Workbook()
    ws_questionario = wb_questionario.active
    ws_questionario.title = "SEZ. II-GEST. FIN. CASSA"
    ws_questionario["D6"] = 2022
    ws_questionario["E6"] = 2023
    ws_questionario["F6"] = 2024
    ws_questionario["B8"] = "Fondo cassa complessivo al 31.12"
    ws_questionario["F8"] = 111
    ws_questionario["B9"] = "di cui cassa vincolata"
    ws_questionario["F9"] = 222
    wb_questionario.save(questionario_path)

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2024,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 1
        assert unresolved == 0
        assert ws_result["B2"].value == "222,00"
    finally:
        wb_result.close()


def test_riscossione_sanzioni_codice_strada_uses_questionario_dati_entrate(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2024
    ws_analysis["A2"] = "% RISCOSSIONI SANZIONI DEL CODICE DELLA STRADA"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rend. 2024.xlsx"
    Workbook().save(bdap_path)

    questionario_path = tmp_path / "Questionario Bilancio Enti locali Consuntivo 2024.xlsx"
    wb_questionario = Workbook()
    ws_questionario = wb_questionario.active
    ws_questionario.title = "SEZ. II - DATI ENTRATE_"
    ws_questionario["B5"] = "Residui attivi"
    ws_questionario["E5"] = 2020
    ws_questionario["F5"] = 2021
    ws_questionario["G5"] = 2022
    ws_questionario["H5"] = 2023
    ws_questionario["I5"] = 2024
    ws_questionario["B18"] = "Sanzioni per violazioni Codice della strada"
    ws_questionario["C18"] = "Residui iniziali"
    ws_questionario["C19"] = "Riscosso c/residui al 31.12"
    ws_questionario["C20"] = "Percentuale di riscossione"
    ws_questionario["H20"] = "85,54%"
    ws_questionario["I20"] = "100%"
    wb_questionario.save(questionario_path)

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2024,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 1
        assert unresolved == 0
        assert ws_result["B2"].value == "100,00%"
        assert "SEZ. II - DATI ENTRATE_" in ws_result["B2"].comment.text
        assert "I20" in ws_result["B2"].comment.text
    finally:
        wb_result.close()


def test_riscossione_sanzioni_codice_strada_uses_2021_gest_entrate_layout(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2021
    ws_analysis["A2"] = "% RISCOSSIONI SANZIONI DEL CODICE DELLA STRADA"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "2021_rend_ind.xlsx"
    Workbook().save(bdap_path)

    questionario_path = tmp_path / "Questionario Bilancio Enti locali Consuntivo 2021.xlsx"
    wb_questionario = Workbook()
    ws_questionario = wb_questionario.active
    ws_questionario.title = "SEZ. II- Gest. ENTRATE"
    ws_questionario["A23"] = "Residui attivi"
    ws_questionario["C23"] = "Esercizi precedenti"
    ws_questionario["D23"] = 2017
    ws_questionario["E23"] = 2018
    ws_questionario["F23"] = 2019
    ws_questionario["G23"] = 2020
    ws_questionario["H23"] = 2021
    ws_questionario["A30"] = "Sanzioni per violazioni codice della strada"
    ws_questionario["B30"] = "Residui iniziali"
    ws_questionario["B31"] = "Riscosso c/residui al 31.12"
    ws_questionario["B32"] = "Percentuale di riscossione"
    ws_questionario["H32"] = "100%"
    wb_questionario.save(questionario_path)

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2021,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 1
        assert unresolved == 0
        assert ws_result["B2"].value == "100,00%"
        assert "SEZ. II- Gest. ENTRATE" in ws_result["B2"].comment.text
        assert "H32" in ws_result["B2"].comment.text
    finally:
        wb_result.close()


def test_active_questionario_rows_remain_empty_without_placeholder(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2024
    ws_analysis["A2"] = "CONTROLLO SULLE SOCIETÀ PARTECIPATE"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rend. 2024.xlsx"
    Workbook().save(bdap_path)

    questionario_path = tmp_path / "Questionario 2024.xlsx"
    Workbook().save(questionario_path)

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2024,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 0
        assert unresolved == 0
        assert ws_result["B2"].value is None
        assert ws_result["B2"].comment is None
    finally:
        wb_result.close()
