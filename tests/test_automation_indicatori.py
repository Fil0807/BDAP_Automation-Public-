from openpyxl import Workbook

from bdap_app.core.automation_indicatori import process_indicatori_row


def test_process_indicatori_row_reads_explicit_cell(tmp_path) -> None:
    bdap_path = tmp_path / "Rend. 2024.xlsx"
    Workbook().save(bdap_path)

    indicatori_path = tmp_path / "Indicatori 2024.xlsx"
    wb_indicatori = Workbook()
    ws_indicatori = wb_indicatori.active
    ws_indicatori.title = "Indicatori"
    ws_indicatori["D27"] = 0.25
    wb_indicatori.save(indicatori_path)

    wb_output = Workbook()
    target_cell = wb_output.active["A1"]
    handled, delta_rows, loaded_wb, _, _, resolved = process_indicatori_row(
        target_cell,
        bdap_path,
        2024,
        "incidenzaspesapersonalesuspesacorrente",
        {
            "cell_ref": "D27",
            "is_percentage": True,
            "sheet_idx": "1",
            "source_workbook": "indicatori",
        },
    )

    try:
        assert handled is True
        assert delta_rows == 1
        assert resolved == 0.25
        assert target_cell.value == "25,00%"
        assert target_cell.comment is not None
        assert "Indicatori 2024.xlsx" in target_cell.comment.text
    finally:
        if loaded_wb is not None:
            loaded_wb.close()


def test_process_indicatori_row_can_use_bdap_path_itself(tmp_path) -> None:
    bdap_path = tmp_path / "2024_rend_ind_sintetici.xlsx"
    wb_indicatori = Workbook()
    ws_indicatori = wb_indicatori.active
    ws_indicatori.title = "Indicatori"
    ws_indicatori["D27"] = 0.33
    wb_indicatori.save(bdap_path)

    wb_output = Workbook()
    target_cell = wb_output.active["A1"]
    handled, delta_rows, loaded_wb, _, loaded_path, resolved = process_indicatori_row(
        target_cell,
        bdap_path,
        2024,
        "incidenzaspesapersonalesuspesacorrente",
        {
            "cell_ref": "D27",
            "is_percentage": True,
            "sheet_idx": "1",
            "source_workbook": "indicatori",
        },
    )

    try:
        assert handled is True
        assert delta_rows == 1
        assert resolved == 0.33
        assert loaded_path == bdap_path
        assert target_cell.value == "33,00%"
        assert target_cell.comment is not None
        assert "2024_rend_ind_sintetici.xlsx" in target_cell.comment.text
    finally:
        if loaded_wb is not None:
            loaded_wb.close()
