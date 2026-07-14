from openpyxl import Workbook, load_workbook

from bdap_app.core.automation import fill_accounting_analysis


def _save_questionario_debiti(path, year: int, answer: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sezione preliminare"
    ws["B19"] = f"4. Sono stati riconosciuti debiti fuori bilancio nel {year}?"
    ws["P19"] = answer
    wb.save(path)


def test_debiti_fuori_bilancio_uses_questionario_for_target_year(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2021
    ws_analysis["A2"] = "Debiti fuori bilancio"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rendiconto 2021.xlsx"
    Workbook().save(bdap_path)

    _save_questionario_debiti(
        tmp_path / "Questionario Debiti Fuori Bilancio Periodico 2024.xlsx",
        2024,
        "Si",
    )
    _save_questionario_debiti(
        tmp_path / "Questionario Debiti Fuori Bilancio Periodico 2021.xlsx",
        2021,
        "No",
    )

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2021,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 1
        assert unresolved == 0
        assert ws_result["B2"].value == "No"
        assert ws_result["B2"].comment is not None
        assert "Questionario Debiti Fuori Bilancio Periodico 2021.xlsx" in ws_result["B2"].comment.text
    finally:
        wb_result.close()


def test_debiti_fuori_bilancio_normalizes_si(tmp_path) -> None:
    analysis_path = tmp_path / "analysis.xlsx"
    wb_analysis = Workbook()
    ws_analysis = wb_analysis.active
    ws_analysis.title = "analisi contabile"
    ws_analysis["A1"] = "Voce"
    ws_analysis["B1"] = 2022
    ws_analysis["A2"] = "debitifuoribilancio"
    wb_analysis.save(analysis_path)

    bdap_path = tmp_path / "Rendiconto 2022.xlsx"
    Workbook().save(bdap_path)
    _save_questionario_debiti(
        tmp_path / "Questionario Debiti Fuori Bilancio Periodico 2022.xlsx",
        2022,
        "SI",
    )

    updated, unresolved = fill_accounting_analysis(
        analysis_path=analysis_path,
        bdap_path=bdap_path,
        output_path=analysis_path,
        year=2022,
        analysis_sheet="analisi contabile",
    )

    wb_result = load_workbook(analysis_path)
    try:
        ws_result = wb_result["analisi contabile"]
        assert updated == 1
        assert unresolved == 0
        assert ws_result["B2"].value == "Sì"
    finally:
        wb_result.close()
