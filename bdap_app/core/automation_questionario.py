"""Helper per la scoperta del workbook `questionario`."""

import logging
import re
from pathlib import Path
from typing import Optional
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils.cell import column_index_from_string
from ..support.text_utils import (
    normalize_text,
    _coerce_year_value,
    _coerce_year_value_importocassa,
    text_options,
)
from ..support.value_resolver import (
    find_bdap_value_by_label,
    find_bdap_value_with_fallback,
    find_bdap_value_by_expected_label_near_cell,
)
from ..support.value_formatter import format_value_italian, coerce_numeric
from ..support.text_utils import _coerce_int_or_default, parse_bool_flag
from ..support.default_mappings import TEMPLATE_SOURCES
from .automation_search import find_nearby_title_row, find_year_column_near_row
from decimal import Decimal
from openpyxl.utils import get_column_letter, coordinate_to_tuple

logger = logging.getLogger(__name__)

QUESTIONARIO_ACTIVE_SPECIAL_KEYS = {
    'fondocassadaidatisiope',
    'spesadelpersonalesostenuta',
    '%riscossionisanzionidelcodicedellastrada',
    '%riscossioniproventidapermessoacostruire',
    'contrastoallevasionetributaria',
    '%diriscossionecomplessiva',
    'disavanzodariaccertamentostraordinario',
    'controllosullesocietàpartecipate',
    'partenariatopubblicoprivato',
    'prospettocoperturaserviziadomanda',
    'resacontoagenti'
}
def _resolve_questionario_year_column(ws_bdap, target_year: int, base_row: int, max_search_rows: int = 20, start_col: int = 1) -> Optional[int]:
    """Trova la colonna dell'anno nel questionario cercando verso l'alto dalla riga base."""
    return find_year_column_near_row(
        ws_bdap,
        target_year,
        base_row,
        max_search_rows=max_search_rows,
        start_col=start_col,
        year_parser=_coerce_year_value_importocassa,
    )


def _resolve_questionario_formula_value(wb_formula, ws_formula, cell_ref: str):
    """Risolve il valore per celle formula che fanno riferimento a DATI tramite SCARTO(...)."""
    try:
        formula_value = ws_formula[cell_ref].value
    except Exception:
        return None

    from .automation import _resolve_scarto_formula_reference

    ref = _resolve_scarto_formula_reference(formula_value)
    if ref is None:
        return None

    sheet_name, row_idx, col_idx = ref
    if sheet_name not in wb_formula.sheetnames:
        return None
    ws_data = wb_formula[sheet_name]
    return ws_data.cell(row=row_idx, column=col_idx).value


def _clear_questionario_manual_cell(target_cell) -> None:
    """Lascia vuote le righe del questionario non compilabili automaticamente."""
    target_cell.value = None
    target_cell.comment = None


def _resolve_questionario_sheet(source_wb, mapping: dict):
    """Risolve il foglio del questionario usando uno o più token oppure l'indice."""
    for token in text_options(mapping.get("sheet_name_contains")):
        token_norm = normalize_text(token)
        for name in source_wb.sheetnames:
            if token_norm in normalize_text(name):
                return source_wb[name]

    sheet_idx = mapping.get("sheet_idx")
    if sheet_idx is not None:
        try:
            idx = int(float(str(sheet_idx).strip()))
        except Exception:
            idx = None
        if idx is not None and 1 <= idx <= len(source_wb.sheetnames):
            return source_wb[source_wb.sheetnames[idx - 1]]

    return None


def _questionario_table_refs(mapping: dict) -> list[tuple[str, object]]:
    """Restituisce coppie (cella valore, cella/e intestazione anni) per tabelle alternative."""
    explicit_refs = mapping.get("table_refs")
    if isinstance(explicit_refs, (list, tuple)):
        pairs: list[tuple[str, object]] = []
        for item in explicit_refs:
            if not isinstance(item, dict):
                continue
            cell_ref = item.get("cell_ref")
            year_ref = item.get("year_cell_ref")
            if isinstance(cell_ref, str) and cell_ref.strip() and year_ref:
                pairs.append((cell_ref.strip(), year_ref))
        if pairs:
            return pairs

    cell_ref = mapping.get("cell_ref")
    year_ref = mapping.get("year_cell_ref")
    if isinstance(cell_ref, str) and cell_ref.strip() and year_ref:
        return [(cell_ref.strip(), year_ref)]
    return []


def _resolve_fallback_display_ref(source_wb, formula_wb, year: int, fallback_mapping: dict):
    ws_source = _resolve_questionario_sheet(source_wb, fallback_mapping)
    if ws_source is None:
        return None, None, None, None

    cell_ref = fallback_mapping.get("cell_ref")
    if not isinstance(cell_ref, str) or not cell_ref.strip():
        return None, None, None, ws_source.title

    row_tol = max(0, _coerce_int_or_default(fallback_mapping.get("row_tolerance", 2), 2))
    try:
        base_row, base_col = coordinate_to_tuple(cell_ref)
    except Exception:
        return None, None, None, ws_source.title

    target_row = base_row
    expected_label = fallback_mapping.get("expected_label")
    if expected_label:
        found_title_row = find_nearby_title_row(
            ws_source,
            str(expected_label),
            base_row,
            base_col,
            row_tolerance=row_tol,
        )
        if found_title_row is None:
            expected_norm = normalize_text(expected_label)
            row_start = max(1, base_row - row_tol)
            row_end = min(ws_source.max_row, base_row + row_tol)
            col_end = min(ws_source.max_column, base_col)
            for row_idx in range(row_start, row_end + 1):
                for col_idx in range(1, col_end + 1):
                    cell_value = ws_source.cell(row=row_idx, column=col_idx).value
                    if isinstance(cell_value, str) and expected_norm in normalize_text(cell_value):
                        found_title_row = row_idx
                        break
                if found_title_row is not None:
                    break
        if found_title_row is None:
            return None, None, None, ws_source.title
        target_row = found_title_row

    actual_ref = f"{get_column_letter(base_col)}{target_row}"
    candidate, candidate_ref = find_bdap_value_with_fallback(ws_source, actual_ref, max_offset=0)
    candidate_num = coerce_numeric(candidate)
    if candidate_num in (None, 0) and formula_wb is not None and ws_source.title in formula_wb.sheetnames:
        formula_value = _resolve_questionario_formula_value(
            formula_wb,
            formula_wb[ws_source.title],
            candidate_ref or actual_ref,
        )
        formula_num = coerce_numeric(formula_value)
        if formula_num not in (None, 0):
            candidate = formula_value
            candidate_num = formula_num

    return candidate, candidate_num, candidate_ref or actual_ref, ws_source.title


def _discover_questionario_workbook(bdap_path: Optional[Path], year: int, special_key: Optional[str] = None) -> Optional[Path]:
    """Localizza il workbook 'questionario' per l'anno richiesto nella stessa directory del BDAP."""
    if bdap_path is None:
        logger.debug("_discover_questionario_workbook: bdap_path is None")
        return None
    base_dir = bdap_path.parent
    logger.debug(f"_discover_questionario_workbook: searching in {base_dir} for year={year}, special_key={special_key}")
    if not base_dir.exists() or not base_dir.is_dir():
        logger.debug(f"_discover_questionario_workbook: base_dir does not exist or is not a dir: {base_dir}")
        return None

    def _extract_years(path: Path) -> set[int]:
        text = f"{path.name} {path.stem}"
        return {int(m) for m in re.findall(r"(19\d{2}|20\d{2})", text)}

    def _filename_year_priority(path: Path) -> tuple[int, int, str]:
        years_in_name = sorted(_extract_years(path))
        if year in years_in_name:
            return 0, year, path.name.lower()

        later_years = [candidate_year for candidate_year in years_in_name if candidate_year > year]
        if later_years:
            return 1, min(later_years), path.name.lower()

        earlier_years = [candidate_year for candidate_year in years_in_name if candidate_year < year]
        if earlier_years:
            return 2, max(earlier_years), path.name.lower()

        return 3, 9999, path.name.lower()

    def _candidate_table_years(path: Path) -> set[int]:
        """Legge gli anni dalla riga di intestazione indicata dalla mappatura."""
        if not special_key or special_key not in TEMPLATE_SOURCES:
            return set()

        mapping = TEMPLATE_SOURCES[special_key]
        table_refs = _questionario_table_refs(mapping)
        if not table_refs:
            return set()

        header_rows: set[int] = set()
        for _, year_ref in table_refs:
            refs = year_ref if isinstance(year_ref, (list, tuple)) else [year_ref]
            for cell_ref in refs:
                try:
                    row_idx, _ = coordinate_to_tuple(str(cell_ref))
                    header_rows.add(row_idx)
                except Exception:
                    continue
        if not header_rows:
            return set()

        row_tol = max(0, _coerce_int_or_default(mapping.get("row_tolerance", 1), 1))
        years: set[int] = set()
        try:
            wb_temp = load_workbook(path, read_only=True, data_only=True)
            ws_temp = _resolve_questionario_sheet(wb_temp, mapping) or wb_temp.active
            max_col = ws_temp.max_column or 12
            rows_to_scan: set[int] = set()
            for row_idx in header_rows:
                for offset in range(-row_tol, row_tol + 1):
                    candidate_row = row_idx + offset
                    if candidate_row >= 1:
                        rows_to_scan.add(candidate_row)

            for row_idx in sorted(rows_to_scan):
                for col_idx in range(1, max_col + 1):
                    year_value = _coerce_year_value_importocassa(
                        ws_temp.cell(row=row_idx, column=col_idx).value
                    )
                    if year_value is not None:
                        years.add(year_value)
            wb_temp.close()
        except Exception:
            return set()
        return years

    prefer_latest_table = False
    prefer_exact_table = False
    if special_key and special_key in TEMPLATE_SOURCES:
        source_rule = TEMPLATE_SOURCES[special_key]
        prefer_latest_table = parse_bool_flag(source_rule.get("prefer_latest_table_workbook"), default=False)
        prefer_exact_table = parse_bool_flag(source_rule.get("prefer_exact_table_workbook"), default=False)

    table_years_cache: dict[Path, set[int]] = {}

    def _cached_table_years(path: Path) -> set[int]:
        if path not in table_years_cache:
            table_years_cache[path] = _candidate_table_years(path)
        return table_years_cache[path]

    def _candidate_sort_key(path: Path) -> tuple:
        filename_priority = _filename_year_priority(path)
        root_priority = candidates_by_path[path]
        if filename_priority[0] == 0:
            return (0, root_priority, *filename_priority)

        if prefer_exact_table:
            table_years = _cached_table_years(path)
            if year in table_years:
                return (1, root_priority, 0, *filename_priority)
            return (1, root_priority, 1, *filename_priority)

        if prefer_latest_table:
            table_years = _cached_table_years(path)
            if year in table_years:
                latest_table_year = max(table_years) if table_years else 0
                filename_years = _extract_years(path)
                latest_filename_year = max(filename_years) if filename_years else 0
                return (
                    1,
                    root_priority,
                    0,
                    -latest_table_year,
                    -latest_filename_year,
                    path.name.lower(),
                )
            return (1, root_priority, 1, *filename_priority)

        return (1, root_priority, *filename_priority)

    def _candidate_year_matches(path: Path) -> bool:
        """Valida l'anno target dalle celle specifiche per special_key, oppure dal nome file."""
        try:
            wb_temp = load_workbook(path, read_only=True, data_only=True)
            found_year = False
            ws_temp = wb_temp.active
            
            # Se special_key ha celle specifiche, controlla lì
            if special_key and special_key in TEMPLATE_SOURCES:
                mapping = TEMPLATE_SOURCES[special_key]
                ws_temp = _resolve_questionario_sheet(wb_temp, mapping) or ws_temp

                table_refs = _questionario_table_refs(mapping)
                if table_refs:
                    year_parser = _coerce_year_value_importocassa if special_key == "importocassavincolata" else _coerce_year_value
                    for _, year_ref in table_refs:
                        year_cells = year_ref if isinstance(year_ref, (list, tuple)) else [year_ref]
                        for cell_ref in year_cells:
                            try:
                                row, col = coordinate_to_tuple(str(cell_ref))
                                val = ws_temp.cell(row=row, column=col).value
                                if year_parser(val) == year:
                                    found_year = True
                                    break
                            except Exception:
                                pass
                        if found_year:
                            break
            
            # Se non trovato con year_cell_ref, scansione standard nelle prime 10 righe
            if not found_year:
                for r in range(1, 11):
                    # ws_temp.max_column can be None in read_only mode; fall back to 12
                    try:
                        max_col = ws_temp.max_column if ws_temp.max_column is not None else 12
                    except Exception:
                        max_col = 12
                    for c in range(1, min(max_col, 12) + 1):
                        try:
                            val = ws_temp.cell(row=r, column=c).value
                            if _coerce_year_value(val) == year:
                                found_year = True
                                break
                        except Exception:
                            pass
                    if found_year:
                        break
            
            wb_temp.close()
            if found_year:
                logger.debug(f"_candidate_year_matches: {path} has year {year} in cells for special_key={special_key}")
                return True
            
            years_in_name = _extract_years(path)
            logger.debug(f"_candidate_year_matches: {path} has years in name: {years_in_name}, special_key={special_key}")
            return year in years_in_name
        except Exception as e:
            logger.debug(f"_candidate_year_matches: Exception checking {path}: {e}")
            return False

    def _is_questionario_candidate(path: Path) -> bool:
        """Determina se un file è un candidato valido per essere il questionario, escludendo BDAP e file temporanei."""
        if not path.is_file() or path.suffix.lower() not in {".xlsx", ".xlsm"}:
            return False
        if path == bdap_path:
            return False
        if path.name.startswith("~$"):
            return False
        normalized_name = normalize_text(path.name)
        if "controllipost" in normalized_name:
            logger.debug(f"_is_questionario_candidate: {path} is a Controlli Post workbook")
            return False
        is_candidate = "questionario" in normalize_text(str(path))
        if not is_candidate:
            logger.debug(f"_is_questionario_candidate: {path} does not contain 'questionario'")
        return is_candidate

    # Determina le root directories da cercare
    if special_key == "importocassavincolata":
        search_roots: list[Path] = [base_dir]
        if base_dir.parent.exists() and base_dir.parent.is_dir():
            search_roots.append(base_dir.parent)
        if base_dir.parent.parent.exists() and base_dir.parent.parent.is_dir():
            search_roots.append(base_dir.parent.parent)

        # Preserva l'ordine rimuovendo i duplicati.
        unique_roots: list[Path] = []
        seen_roots: set[Path] = set()
        for root in search_roots:
            try:
                resolved = root.resolve()
            except Exception:
                resolved = root
            if resolved in seen_roots:
                continue
            seen_roots.add(resolved)
            unique_roots.append(root)
    else:
        # Start from the BDAP directory and expand only to nearby parent folders.
        candidate_roots: list[Path] = [base_dir]
        try:
            if base_dir.parent.exists() and base_dir.parent.is_dir():
                candidate_roots.append(base_dir.parent)
            if base_dir.parent.parent.exists() and base_dir.parent.parent.is_dir():
                candidate_roots.append(base_dir.parent.parent)
        except Exception:
            pass

        # If the BDAP is in a folder named 'datibdap', also consider its parent
        try:
            if normalize_text(base_dir.name) == "datibdap" and base_dir.parent.exists() and base_dir.parent.is_dir():
                candidate_roots.append(base_dir.parent)
        except Exception:
            pass

        # Preserve order and remove duplicates
        unique_roots = []
        seen_roots: set[Path] = set()
        for root in candidate_roots:
            try:
                resolved = root.resolve()
            except Exception:
                resolved = root
            if resolved in seen_roots:
                continue
            seen_roots.add(resolved)
            unique_roots.append(root)
    
    logger.debug(f"_discover_questionario_workbook: search_roots = {unique_roots}")
    
    candidates_by_path: dict[Path, int] = {}
    for root_priority, root in enumerate(unique_roots):
        for pattern in ("*.xlsx", "*.xlsm"):
            for candidate in root.rglob(pattern):
                if _is_questionario_candidate(candidate):
                    current_priority = candidates_by_path.get(candidate)
                    if current_priority is None or root_priority < current_priority:
                        candidates_by_path[candidate] = root_priority
    
    logger.debug(f"_discover_questionario_workbook: candidates matching 'questionario': {set(candidates_by_path)}")
    
    candidates = sorted(candidates_by_path, key=_candidate_sort_key)
    if not candidates:
        logger.debug(f"_discover_questionario_workbook: no candidates found")
        return None

    
    for candidate in candidates:
        if year in _extract_years(candidate) and _candidate_year_matches(candidate):
            logger.debug(f"_discover_questionario_workbook: selected {candidate}")
            return candidate

    logger.debug(f"_discover_questionario_workbook: no candidate filename contains target year {year}")
    return None


def _load_questionario_workbooks(
    bdap_path: Optional[Path],
    year: int,
    special_key: Optional[str] = None,
    workbook_cache: Optional[dict[str, tuple[object, object]]] = None,
):
    """Carica (e memorizza in cache) i workbook questionario (data-only e formula) se presenti.

    Restituisce una tupla `(wb_data, wb_formula, path)` o `(None, None, None)` se non trovati.
    """
    if bdap_path is None:
        return None, None, None

    try:
        path = _discover_questionario_workbook(bdap_path, int(year), special_key=special_key)
        if path is None:
            return None, None, None

        cache_key = str(path)
        if workbook_cache is not None and cache_key in workbook_cache:
            wb_data, wb_formula = workbook_cache[cache_key]
            return wb_data, wb_formula, path

        try:
            wb_data = load_workbook(path, data_only=True)
            wb_formula = load_workbook(path, data_only=False)
        except Exception:
            return None, None, None

        if workbook_cache is not None:
            workbook_cache[cache_key] = (wb_data, wb_formula)
        return wb_data, wb_formula, path
    except Exception:
        return None, None, None


def process_questionario_row(
    target_cell,
    bdap_path: Optional[Path],
    year: int,
    special_key: Optional[str],
    data_source_mapping: dict,
    keep_source_reference: bool = True,
    workbook_cache: Optional[dict[str, tuple[object, object]]] = None,
):
    """Handle full resolution for a `questionario` source.

    Returns (handled: bool, delta_rows: int, wb_questionario, wb_questionario_formula, questionario_path_cache, resolved_numeric_tuple|None)
    """
    # Load cached workbooks
    wb_questionario, wb_questionario_formula, questionario_path_cache = _load_questionario_workbooks(
        bdap_path, year, special_key, workbook_cache=workbook_cache
    )

    # No workbook found
    if wb_questionario is None:
        if special_key in QUESTIONARIO_ACTIVE_SPECIAL_KEYS:
            _clear_questionario_manual_cell(target_cell)
            return True, 0, None, None, None, None
        if data_source_mapping.get("optional", False):
            target_cell.value = None
            target_cell.comment = None
            return True, 0, None, None, None, None
        raise KeyError("Missing questionario workbook")

    source_wb = wb_questionario
    questionario_formula_wb = wb_questionario_formula
    source_file_name = questionario_path_cache.name if questionario_path_cache is not None else (bdap_path.name if bdap_path is not None else "n/d")

    # For contrazionenuovimutui prefer questionario files that include the target year in the filename.
    if special_key == "contrazionenuovimutui" and questionario_path_cache is not None:
        try:
            years_in_name = {int(m) for m in re.findall(r"(19\d{2}|20\d{2})", questionario_path_cache.name)}
        except Exception:
            years_in_name = set()
        if year not in years_in_name:
            logger.warning(
                "Questionario file %s does not include year %s in filename; will skip this cell",
                questionario_path_cache.name,
                year,
            )
            # Continua la risoluzione di altre celle 
            return False, 0, wb_questionario, questionario_formula_wb, questionario_path_cache, None

    # Resolve sheet by name contains / idx
    ws_bdap_src = _resolve_questionario_sheet(source_wb, data_source_mapping)
    has_display_fallback = bool(data_source_mapping.get("fallback_display_refs")) and "display_refs" in data_source_mapping

    if ws_bdap_src is None and not has_display_fallback:
        if special_key in QUESTIONARIO_ACTIVE_SPECIAL_KEYS:
            _clear_questionario_manual_cell(target_cell)
            return True, 0, wb_questionario, questionario_formula_wb, questionario_path_cache, None
        if data_source_mapping.get("optional", False):
            target_cell.value = None
            target_cell.comment = None
            return True, 0, wb_questionario, questionario_formula_wb, questionario_path_cache, None
        # No sheet found -> let caller treat as not handled
        return False, 0, wb_questionario, questionario_formula_wb, questionario_path_cache, None

    # Value extraction (subset of logic to mirror automation.py behavior)
    value = None
    actual_cell_ref = None

    if (
        special_key in QUESTIONARIO_ACTIVE_SPECIAL_KEYS
        and not any(key in data_source_mapping for key in ("first_non_zero", "display_refs", "cell_ref", "label"))
    ):
        _clear_questionario_manual_cell(target_cell)
        return True, 0, wb_questionario, questionario_formula_wb, questionario_path_cache, None

    if "first_non_zero" in data_source_mapping:
        row_tol = max(0, _coerce_int_or_default(data_source_mapping.get("row_tolerance", 2), 2))
        for option in data_source_mapping["first_non_zero"]:
            ws_opt = None
            try:
                ws_opt = source_wb[source_wb.sheetnames[int(option["sheet_idx"]) - 1]]
            except Exception:
                ws_opt = None
            if ws_opt is None:
                continue
            candidate, candidate_ref = find_bdap_value_with_fallback(ws_opt, option["cell_ref"], max_offset=row_tol)
            if coerce_numeric(candidate) not in (None, 0):
                value = candidate
                ws_bdap_src = ws_opt
                actual_cell_ref = candidate_ref
                break
        if value is None:
            option = data_source_mapping["first_non_zero"][0]
            ws_bdap_src = source_wb[source_wb.sheetnames[int(option["sheet_idx"]) - 1]]
            value, actual_cell_ref = find_bdap_value_with_fallback(ws_bdap_src, option["cell_ref"], max_offset=row_tol)

    elif "display_refs" in data_source_mapping:
        # Handle cases like residui where multiple titled cells need to be displayed
        display_refs = data_source_mapping["display_refs"]
        use_static_refs = parse_bool_flag(data_source_mapping.get("static_column"), default=False)
        row_tol = max(0, _coerce_int_or_default(data_source_mapping.get("row_tolerance", 2), 2))
        lines = []
        last_refs = []
        source_descriptions = []
        for title, cell_ref in display_refs:
            try:
                base_row, base_col = coordinate_to_tuple(cell_ref)
            except Exception:
                base_row, base_col = None, None

            candidate = None
            candidate_ref = None
            year_aborted = False
            source_sheet_title = ws_bdap_src.title if ws_bdap_src is not None else None

            if base_row is not None and ws_bdap_src is not None:
                target_row = base_row
                found_title_row = find_nearby_title_row(
                    ws_bdap_src,
                    str(title),
                    base_row,
                    base_col,
                    row_tolerance=row_tol,
                )
                if found_title_row is not None:
                    target_row = found_title_row

                y_col = None
                if not use_static_refs and "year_cell_ref" in data_source_mapping:
                    refs = data_source_mapping["year_cell_ref"]
                    if isinstance(refs, (list, tuple)):
                        for yref in refs:
                            try:
                                y_row, y_col_idx = coordinate_to_tuple(yref)
                                val = ws_bdap_src.cell(row=y_row, column=y_col_idx).value
                                if _coerce_year_value(val) == year:
                                    y_col = y_col_idx
                                    break
                            except Exception:
                                continue
                    else:
                        try:
                            y_row, y_col_idx = coordinate_to_tuple(refs)
                            val = ws_bdap_src.cell(row=y_row, column=y_col_idx).value
                            if _coerce_year_value(val) == year:
                                y_col = y_col_idx
                        except Exception:
                            y_col = None

                if not use_static_refs and y_col is None:
                    # try resolving by scanning nearby header rows
                    try:
                        y_col = _resolve_questionario_year_column(ws_bdap_src, year, target_row, start_col=4)
                    except Exception:
                        y_col = None

                if y_col is not None:
                    year_cell = ws_bdap_src.cell(row=target_row, column=y_col)
                    candidate = year_cell.value
                    candidate_ref = year_cell.coordinate
                else:
                    # Legge il riferimento statico quando la mappatura lo richiede
                    # oppure quando non è stata risolta una colonna anno.
                    actual_ref = f"{get_column_letter(base_col)}{target_row}" if base_col else cell_ref
                    candidate, candidate_ref = find_bdap_value_with_fallback(ws_bdap_src, actual_ref, max_offset=0)

            candidate_num = coerce_numeric(candidate)
            if candidate_num in (None, 0) and questionario_formula_wb is not None:
                ws_formula_src = None
                if ws_bdap_src is not None and ws_bdap_src.title in questionario_formula_wb.sheetnames:
                    ws_formula_src = questionario_formula_wb[ws_bdap_src.title]
                if ws_formula_src is not None and candidate_ref:
                    formula_value = _resolve_questionario_formula_value(
                        questionario_formula_wb,
                        ws_formula_src,
                        candidate_ref,
                    )
                    formula_num = coerce_numeric(formula_value)
                    if formula_num not in (None, 0):
                        candidate = formula_value
                        candidate_num = formula_num

            if candidate_num in (None, 0):
                fallback_refs = data_source_mapping.get("fallback_display_refs", [])
                if isinstance(fallback_refs, dict):
                    fallback_refs = [fallback_refs]
                for fallback_ref in fallback_refs:
                    if not isinstance(fallback_ref, dict):
                        continue
                    fallback_title = fallback_ref.get("title")
                    if fallback_title and normalize_text(fallback_title) != normalize_text(title):
                        continue
                    fallback_value, fallback_num, fallback_cell_ref, fallback_sheet_title = _resolve_fallback_display_ref(
                        source_wb,
                        questionario_formula_wb,
                        year,
                        fallback_ref,
                    )
                    if fallback_num not in (None, 0):
                        candidate = fallback_value
                        candidate_num = fallback_num
                        candidate_ref = fallback_cell_ref
                        source_sheet_title = fallback_sheet_title
                        break

            coerced = candidate_num
            if coerced is None:
                num_val = Decimal('0')
                last_refs.append(cell_ref)
                source_descriptions.append(f"{title}: cella {cell_ref}")
            else:
                num_val = coerced
                last_refs.append(candidate_ref or cell_ref)
                if source_sheet_title:
                    source_descriptions.append(f"{title}: foglio {source_sheet_title}, cella {candidate_ref or cell_ref}")
                else:
                    source_descriptions.append(f"{title}: cella {candidate_ref or cell_ref}")

            formatted = format_value_italian(num_val)
            lines.append(f"{title}: {formatted}")

        target_cell.value = "\n".join(lines)
        try:
            target_cell.alignment = __import__('openpyxl').styles.Alignment(wrapText=True)
        except Exception:
            pass
        if keep_source_reference and bdap_path is not None:
            if source_descriptions:
                src_descr = " | ".join(source_descriptions)
            else:
                source_refs = last_refs if last_refs else [cell_ref for _, cell_ref in display_refs]
                source_sheet = ws_bdap_src.title if ws_bdap_src is not None else "n/d"
                src_descr = f"foglio {source_sheet} | " + " | ".join(
                    f"{title}: cella {source_ref}"
                    for (title, _), source_ref in zip(display_refs, source_refs)
                )
            target_cell.comment = Comment(f"Fonte BDAP: {source_file_name} | {src_descr}", "BDAP Automation")
        return True, 1, wb_questionario, questionario_formula_wb, questionario_path_cache, None

    elif "cell_ref" in data_source_mapping:
        row_tol = max(0, _coerce_int_or_default(data_source_mapping.get("row_tolerance", 1), 1))
        actual_base_ref = data_source_mapping["cell_ref"]
        fallback_refs = data_source_mapping.get("fallback_cell_refs", [])
        if isinstance(fallback_refs, str):
            fallback_refs = [fallback_refs]
        candidate_base_refs = [actual_base_ref] + [ref for ref in fallback_refs if ref]

        if "year_cell_ref" in data_source_mapping:
            selected_ref = None
            for table_cell_ref, table_year_ref in _questionario_table_refs(data_source_mapping):
                try:
                    v_row, v_col = coordinate_to_tuple(table_cell_ref)
                except Exception:
                    continue

                expected_label = data_source_mapping.get("expected_label")
                if expected_label:
                    found_label_row = find_nearby_title_row(
                        ws_bdap_src,
                        str(expected_label),
                        v_row,
                        v_col,
                        row_tolerance=row_tol,
                    )
                    if found_label_row is None:
                        continue
                    v_row = found_label_row

                year_refs = table_year_ref if isinstance(table_year_ref, (list, tuple)) else [table_year_ref]
                y_col = None
                for year_ref in year_refs:
                    try:
                        y_row, y_start_col = coordinate_to_tuple(str(year_ref))
                    except Exception:
                        continue
                    y_col = _resolve_questionario_year_column(
                        ws_bdap_src,
                        year,
                        y_row,
                        max_search_rows=0,
                        start_col=y_start_col,
                    )
                    if y_col is not None:
                        break

                if y_col is None:
                    continue

                selected_ref = f"{get_column_letter(y_col)}{v_row}"
                break

            if selected_ref is not None:
                actual_base_ref = selected_ref
                candidate_base_refs = [actual_base_ref]
            else:
                target_cell.value = None
                target_cell.comment = None
                if data_source_mapping.get("optional", False):
                    return True, 0, wb_questionario, questionario_formula_wb, questionario_path_cache, None
                return False, 0, wb_questionario, questionario_formula_wb, questionario_path_cache, None

        for base_ref in candidate_base_refs:
            value, actual_cell_ref = find_bdap_value_with_fallback(ws_bdap_src, base_ref, max_offset=row_tol)
            if value is not None:
                break

            # Fallback per "contrazionenuovimutui"
            if special_key == "contrazionenuovimutui" and value is not None:
                try:
                    # Controllo che l'anno sia corretto prima di accettare il valore
                    check_ref = actual_cell_ref or (candidate_base_refs[0] if candidate_base_refs else None)
                    if check_ref:
                        r_idx, _ = coordinate_to_tuple(check_ref)
                        found_year = False
                        # scan first 6 columns of that row for a year matching target
                        for c in range(1, min(ws_bdap_src.max_column or 6, 6) + 1):
                            try:
                                cell_val = ws_bdap_src.cell(row=r_idx, column=c).value
                                if _coerce_year_value(cell_val) == year:
                                    found_year = True
                                    break
                            except Exception:
                                continue
                        if not found_year:
                            logger.warning(f"contrazionenuovimutui: skipping value at row {r_idx} because year {year} not present in row")
                            target_cell.value = None
                            target_cell.comment = None
                            return True, 1, wb_questionario, questionario_formula_wb, questionario_path_cache, None
                except Exception:
                    pass

    else:
        # default: by label
        value, actual_cell_ref = find_bdap_value_by_label(ws_bdap_src, data_source_mapping.get("label"))

    # Fallback formula resolution
    if coerce_numeric(value) in (None, 0) and questionario_formula_wb is not None:
        current_num = coerce_numeric(value)
        if current_num in (None, 0):
            formula_candidates = []
            if isinstance(actual_cell_ref, str) and actual_cell_ref:
                formula_candidates.append(actual_cell_ref)
            primary_ref = data_source_mapping.get("cell_ref")
            if isinstance(primary_ref, str) and primary_ref:
                formula_candidates.append(primary_ref)
            fallback_refs = data_source_mapping.get("fallback_cell_refs", [])
            if isinstance(fallback_refs, str):
                fallback_refs = [fallback_refs]
            for ref in fallback_refs:
                if isinstance(ref, str) and ref:
                    formula_candidates.append(ref)

            ws_formula_src = None
            if ws_bdap_src is not None and ws_bdap_src.title in questionario_formula_wb.sheetnames:
                ws_formula_src = questionario_formula_wb[ws_bdap_src.title]

            if ws_formula_src is not None:
                for formula_cell_ref in formula_candidates:
                    resolved_formula_value = _resolve_questionario_formula_value(
                        questionario_formula_wb,
                        ws_formula_src,
                        formula_cell_ref,
                    )
                    if coerce_numeric(resolved_formula_value) not in (None, 0):
                        value = resolved_formula_value
                        break

    if value is None:
        if data_source_mapping.get("optional", False):
            target_cell.value = None
            target_cell.comment = None
            return False, 0, wb_questionario, questionario_formula_wb, questionario_path_cache, None
        raise KeyError(f"No value found for label {data_source_mapping.get('label')}")

    percent_flag = bool(data_source_mapping.get('is_percentage', False))
    formatted_value = format_value_italian(value, percent=percent_flag)
    target_cell.value = formatted_value
    if keep_source_reference and bdap_path is not None:
        src_descr = f"foglio {ws_bdap_src.title} | cella {actual_cell_ref}"
        if "label" in data_source_mapping:
            src_descr += f" | voce {data_source_mapping['label']}"
        target_cell.comment = Comment(f"Fonte BDAP: {source_file_name} | {src_descr}", "BDAP Automation")

    numeric_value = coerce_numeric(value)
    resolved_numeric = None
    if numeric_value is not None and special_key is not None:
        resolved_numeric = (special_key, numeric_value)

    return True, 1, wb_questionario, questionario_formula_wb, questionario_path_cache, resolved_numeric
