"""Helper di ricerca e risoluzione usati dall'automazione Excel."""

import re
from typing import Callable, Optional, Tuple
from openpyxl.utils.cell import coordinate_to_tuple
from ..support.text_utils import _coerce_year_value, label_match_score, normalize_text


BDAP_REFERENCE_PATTERN = re.compile(
    r"BDAP\s*,?\s*foglio\s*(\d+)\s*,\s*#?\s*([A-Za-z]{1,3}\d+)",
    re.IGNORECASE,
)


def parse_reference(ref_text: str) -> Optional[Tuple[str, str]]:
    """Analizza un riferimento BDAP testuale in (indice_foglio, riferimento_cella)."""
    if not isinstance(ref_text, str):
        return None
    match = BDAP_REFERENCE_PATTERN.search(ref_text)
    if not match:
        return None
    return match.group(1), match.group(2).upper()


def find_sheet(wb, name: str):
    """Trova un worksheet per nome (case-insensitive)."""
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() == name.strip().lower():
            return wb[sheet_name]
    raise ValueError(f"Sheet '{name}' not found. Available sheets: {wb.sheetnames}")


def find_source_mapping(normalized_row_label: str, sources: Optional[dict[str, dict]] = None) -> tuple[Optional[str], Optional[dict]]:
    """Risolve la regola di sorgente speciale e restituisce la chiave e la mappatura della sorgente."""
    if not normalized_row_label:
        return None, None
    mapping_sources = sources or {}
    if normalized_row_label in mapping_sources:
        return normalized_row_label, mapping_sources[normalized_row_label]
    for key, source in mapping_sources.items():
        key_norm = normalize_text(key)
        if key_norm in normalized_row_label or normalized_row_label in key_norm:
            return key, source

    best_key = None
    best_source = None
    best_score = 0.0
    for key, source in mapping_sources.items():
        score = label_match_score(key, normalized_row_label)
        if score > best_score:
            best_score = score
            best_key = key
            best_source = source
    if best_key is not None and best_score >= 0.84:
        return best_key, best_source
    return None, None


def find_reference_column(ws, max_header_rows: int = 10) -> Optional[int]:
    """Restituisce la colonna con il maggior numero di riferimenti BDAP."""
    best_col = None
    best_count = 0
    first_data_row = min(max_header_rows + 1, ws.max_row)
    for col_idx in range(1, ws.max_column + 1):
        count = 0
        for row_idx in range(first_data_row, ws.max_row + 1):
            if parse_reference(ws.cell(row=row_idx, column=col_idx).value):
                count += 1
        if count > best_count:
            best_count = count
            best_col = col_idx
    return best_col


def find_year_column(ws, year: int, max_header_rows: int = 10) -> Optional[int]:
    """Individua la colonna header contenente l'anno scansionando le righe superiori."""
    candidates = {str(year), year, float(year)}
    for row_idx in range(1, max_header_rows + 1):
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if isinstance(val, str):
                val = val.strip()
            if val in candidates:
                return col_idx
    return None


def find_year_column_near_row(
    ws,
    target_year: int,
    base_row: int,
    max_search_rows: int = 20,
    start_col: int = 1,
    year_parser: Callable[[object], Optional[int]] = _coerce_year_value,
) -> Optional[int]:
    """Trova la colonna dell'anno cercando nelle righe sopra una riga di valore."""
    start_row = max(1, base_row - max_search_rows)
    for row_idx in range(start_row, base_row + 1):
        for col_idx in range(start_col, ws.max_column + 1):
            year_value = year_parser(ws.cell(row=row_idx, column=col_idx).value)
            if year_value == target_year:
                return col_idx
    return None


def resolve_sheet_by_contains_or_idx(source_wb, source_mapping: dict, contains_key: str, idx_key: str) -> Optional[object]:
    """Risolve un foglio usando prima un criterio `*_contains`, poi l'indice `*_idx`."""
    if source_wb is None or not isinstance(source_mapping, dict):
        return None

    sheet_name_contains = source_mapping.get(contains_key)
    if isinstance(sheet_name_contains, str) and sheet_name_contains.strip():
        token_norm = normalize_text(sheet_name_contains)
        for name in source_wb.sheetnames:
            if token_norm in normalize_text(name):
                return source_wb[name]

    sheet_idx = source_mapping.get(idx_key)
    if sheet_idx is None:
        return None

    try:
        idx = int(sheet_idx)
    except (ValueError, TypeError):
        try:
            idx_float = float(str(sheet_idx).strip())
        except (ValueError, TypeError):
            return None
        if not idx_float.is_integer():
            return None
        idx = int(idx_float)

    if idx < 1 or idx > len(source_wb.sheetnames):
        return None
    return source_wb[source_wb.sheetnames[idx - 1]]


def find_nearby_title_row(
    ws,
    title: str,
    base_row: int,
    base_col: Optional[int] = None,
    row_tolerance: int = 2,
) -> Optional[int]:
    """Trova una riga con il titolo atteso vicino alla posizione mappata."""
    title_norm = normalize_text(title)
    if not title_norm:
        return None

    row_start = max(1, base_row - row_tolerance)
    row_end = min(ws.max_row, base_row + row_tolerance)
    col_end = min(ws.max_column, base_col if base_col and base_col > 1 else 4)

    for row_idx in range(row_start, row_end + 1):
        for col_idx in range(1, col_end + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if isinstance(cell_value, str) and normalize_text(cell_value) == title_norm:
                return row_idx

    return None


def find_row_by_expected_label(
    ws,
    expected_label: str,
    row_tolerance: int = 10,
    label_check_cells=None,
    search_key_prefix: str = None,
) -> Optional[int]:
    """Trova la riga che contiene l'etichetta attesa, con tolleranza e opzioni di ricerca."""
    if not expected_label:
        return None

    # If search_key_prefix is provided, use it for column A lookup (highest priority)
    if search_key_prefix:
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if isinstance(cell_value, str) and cell_value.strip().startswith(search_key_prefix):
                return row_idx
        return None

    target_norm = normalize_text(expected_label)

    if label_check_cells:
        label_check_cells_list = label_check_cells if isinstance(label_check_cells, list) else [label_check_cells]
        for check_cell_ref in label_check_cells_list:
            try:
                if ':' in str(check_cell_ref):
                    check_cell_ref = str(check_cell_ref).split(':')[0]
                anchor_row, anchor_col = coordinate_to_tuple(check_cell_ref)
            except Exception:
                continue

            row_start = max(1, anchor_row - row_tolerance)
            row_end = min(ws.max_row, anchor_row + row_tolerance)
            col_start = max(1, anchor_col - 2)
            col_end = min(ws.max_column, anchor_col + 2)

            for row_idx in range(row_start, row_end + 1):
                for col_idx in range(col_start, col_end + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if isinstance(cell_value, str) and target_norm in normalize_text(cell_value):
                        return row_idx

    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, min(5, ws.max_column + 1)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if isinstance(cell_value, str) and target_norm in normalize_text(cell_value):
                return row_idx

    return None
