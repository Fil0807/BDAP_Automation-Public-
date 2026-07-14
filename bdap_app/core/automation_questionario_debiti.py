"""Helper per leggere il questionario Debiti Fuori Bilancio."""

import re
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils.cell import coordinate_to_tuple

from ..support.text_utils import _coerce_int_or_default, label_match_score, normalize_text, text_options
from ..support.value_resolver import format_source_location, is_empty_value


def _extract_years(path: Path) -> set[int]:
    text = f"{path.name} {path.stem}"
    return {int(match) for match in re.findall(r"(19\d{2}|20\d{2})", text)}


def _is_questionario_debiti_candidate(path: Path) -> bool:
    text = normalize_text(str(path))
    if "questionario" not in text:
        return False
    return "debitifuoribilancio" in text or (
        "debit" in text and "fuori" in text and "bilancio" in text
    )


def _workbook_mentions_year(path: Path, year: int) -> bool:
    year_text = str(year)
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False

    try:
        for ws in wb.worksheets:
            max_row = min(ws.max_row or 0, 40)
            max_col = min(ws.max_column or 0, 20)
            for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
                for value in row:
                    if value is None:
                        continue
                    if re.search(rf"\b{year_text}\b", str(value)):
                        return True
    finally:
        wb.close()
    return False


def _discover_questionario_debiti_workbook(bdap_path: Optional[Path], year: int) -> Optional[Path]:
    """Localizza il questionario Debiti Fuori Bilancio dell'anno richiesto."""
    if bdap_path is None:
        return None

    base_dir = bdap_path.parent
    if not base_dir.exists() or not base_dir.is_dir():
        return None

    search_roots = [base_dir]
    if normalize_text(base_dir.name) == "datibdap" and base_dir.parent.exists() and base_dir.parent.is_dir():
        search_roots.append(base_dir.parent)

    candidates = sorted(
        {
            candidate
            for root in search_roots
            for candidate in root.rglob("*.xlsx")
            if candidate != bdap_path
            and not candidate.name.startswith("~$")
            and _is_questionario_debiti_candidate(candidate)
        }
    )
    if not candidates:
        return None

    scored: list[tuple[int, str, Path]] = []
    for candidate in candidates:
        if year in _extract_years(candidate):
            scored.append((0, candidate.name.lower(), candidate))
        elif _workbook_mentions_year(candidate, year):
            scored.append((1, candidate.name.lower(), candidate))
        else:
            scored.append((2, candidate.name.lower(), candidate))

    scored.sort()
    if scored[0][0] < 2:
        return scored[0][2]
    if len(scored) == 1:
        return scored[0][2]
    return None


def _resolve_questionario_debiti_sheet(source_wb, mapping: dict):
    for sheet_name in text_options(mapping.get("sheet_name")):
        sheet_norm = normalize_text(sheet_name)
        for name in source_wb.sheetnames:
            if normalize_text(name) == sheet_norm:
                return source_wb[name]

    for token in text_options(mapping.get("sheet_name_contains")):
        token_norm = normalize_text(token)
        for name in source_wb.sheetnames:
            if token_norm in normalize_text(name):
                return source_wb[name]

    sheet_idx = mapping.get("sheet_idx")
    if sheet_idx is not None:
        try:
            idx = int(float(str(sheet_idx).strip()))
        except Exception:
            idx = None
        if idx is not None and 1 <= idx <= len(source_wb.sheetnames):
            return source_wb[source_wb.sheetnames[idx - 1]]

    return None


def _matches_expected_label(expected_label: str, candidate: object, strict: bool) -> bool:
    if not isinstance(candidate, str) or not candidate.strip():
        return False
    if strict:
        return normalize_text(expected_label) == normalize_text(candidate)
    return label_match_score(expected_label, candidate) >= 0.84


def _find_questionario_debiti_row(ws_source, mapping: dict) -> Optional[int]:
    cell_ref = mapping.get("cell_ref")
    expected_label = mapping.get("expected_label")
    if not isinstance(cell_ref, str) or not cell_ref.strip():
        return None

    try:
        base_row, _ = coordinate_to_tuple(cell_ref)
    except Exception:
        return None

    if not expected_label:
        return base_row

    row_tolerance = max(0, _coerce_int_or_default(mapping.get("row_tolerance", 2), 2))
    strict = bool(mapping.get("strict_expected_label", False))
    offsets = [0]
    for offset in range(1, row_tolerance + 1):
        offsets.extend([-offset, offset])

    label_cells = mapping.get("label_check_cells") or []
    if isinstance(label_cells, str):
        label_cells = [label_cells]

    for label_cell_ref in label_cells:
        try:
            if ":" in str(label_cell_ref):
                label_cell_ref = str(label_cell_ref).split(":", 1)[0]
            anchor_row, anchor_col = coordinate_to_tuple(str(label_cell_ref))
        except Exception:
            continue
        for offset in offsets:
            row_idx = anchor_row + offset
            if row_idx < 1 or row_idx > ws_source.max_row:
                continue
            if _matches_expected_label(expected_label, ws_source.cell(row=row_idx, column=anchor_col).value, strict):
                return row_idx

    row_start = max(1, base_row - row_tolerance)
    row_end = min(ws_source.max_row, base_row + row_tolerance)
    for row_idx in range(row_start, row_end + 1):
        for col_idx in range(1, min(ws_source.max_column, 8) + 1):
            if _matches_expected_label(expected_label, ws_source.cell(row=row_idx, column=col_idx).value, strict):
                return row_idx

    return None


def _normalize_yes_no(value: object):
    if isinstance(value, bool):
        return "Sì" if value else "No"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if value == 1:
            return "Sì"
        if value == 0:
            return "No"
    text = str(value).strip()
    normalized = normalize_text(text)
    if normalized in {"si", "s", "yes", "true", "vero", "1"}:
        return "Sì"
    if normalized in {"no", "n", "false", "falso", "0"}:
        return "No"
    return text


def process_questionario_debiti_row(
    target_cell,
    bdap_path: Optional[Path],
    year: int,
    special_key: Optional[str],
    data_source_mapping: dict,
    keep_source_reference: bool = True,
):
    """Risolve una riga con `source_workbook=questionario_debiti`."""
    path = _discover_questionario_debiti_workbook(bdap_path, int(year))
    if path is None:
        if data_source_mapping.get("optional", False):
            target_cell.value = None
            target_cell.comment = None
            return True, 0, None, None, None, None
        raise KeyError("Missing questionario_debiti workbook")

    wb_source = load_workbook(path, data_only=True)
    try:
        ws_source = _resolve_questionario_debiti_sheet(wb_source, data_source_mapping)
        if ws_source is None:
            if data_source_mapping.get("optional", False):
                target_cell.value = None
                target_cell.comment = None
                return True, 0, None, None, path, None
            raise KeyError("Questionario debiti sheet not found")

        try:
            _, value_col = coordinate_to_tuple(data_source_mapping["cell_ref"])
        except Exception as exc:
            raise KeyError("Invalid questionario_debiti cell_ref") from exc

        row_idx = _find_questionario_debiti_row(ws_source, data_source_mapping)
        if row_idx is None:
            if data_source_mapping.get("optional", False):
                target_cell.value = None
                target_cell.comment = None
                return True, 0, None, None, path, None
            raise KeyError("Questionario debiti label not found")

        value_cell = ws_source.cell(row=row_idx, column=value_col)
        if is_empty_value(value_cell.value):
            if data_source_mapping.get("optional", False):
                target_cell.value = None
                target_cell.comment = None
                return True, 0, None, None, path, None
            raise KeyError("Questionario debiti value not found")

        target_cell.value = _normalize_yes_no(value_cell.value)
        if keep_source_reference:
            target_cell.comment = Comment(
                f"Fonte questionario debiti: {path.name} | {format_source_location(ws_source.title, value_cell.coordinate)}",
                "BDAP Automation",
            )
        return True, 1, None, None, path, value_cell.value
    finally:
        wb_source.close()
