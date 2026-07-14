"""Orchestrazione dell'automazione a livello di fogli Excel."""

import logging
import re
from pathlib import Path
from typing import Callable, Optional
from decimal import Decimal
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple, column_index_from_string
from ..support.text_utils import (
    _coerce_int_or_default,
    _coerce_year_value,
    first_text_option,
    label_match_score,
    normalize_text,
    parse_bool_flag,
    text_options,
    tokenize_text,
)
from ..support.value_formatter import format_value_italian, coerce_numeric
from ..support.value_resolver import (
    find_bdap_value_by_label,
    find_bdap_value_by_expected_label_near_cell,
    find_bdap_value_with_fallback,
    format_source_location,
)
from .automation_fcde import (
    populate_fcde_sheet_in_workbook,
    read_fcde_percentage_from_workbook,
    _resolve_fcde_ratio_from_source,
)
from .automation_questionario import (
    process_questionario_row,
)
from .automation_questionario_debiti import process_questionario_debiti_row
from .automation_indicatori import process_indicatori_row
from .automation_relazione import process_relazione_row
from .automation_search import (
    find_reference_column,
    find_nearby_title_row,
    find_row_by_expected_label,
    find_sheet,
    find_year_column,
    find_year_column_near_row,
    parse_reference,
)

logger = logging.getLogger(__name__)

MANUAL_REVIEW_ONLY_LABELS = {
    "fondocassadaidatisiope",
    "spesadelpersonalesostenutarientraneilimitidicui",
    "riscossioniproventidapermessoacostruire",
    "contrastoallevasionetributaria",
}

DELEGATED_SOURCE_HANDLERS = {
    "questionario": process_questionario_row,
    "questionariodebiti": process_questionario_debiti_row,
    "indicatori": process_indicatori_row,
}


def extract_row_label(ws, row_idx: int, max_columns: int = 4) -> str:
    """Estrae l'etichetta di riga più probabile dal lato sinistro del foglio."""
    for col_idx in range(1, min(max_columns, ws.max_column) + 1):
        value = ws.cell(row=row_idx, column=col_idx).value
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


def resolve_bdap_sheet(wb_bdap, sheet_idx: str):
    """Risolve il foglio BDAP dato un indice 1-based dello sheet."""
    try:
        idx = int(sheet_idx)
    except (ValueError, TypeError):
        try:
            idx_float = float(str(sheet_idx).strip())
        except (ValueError, TypeError):
            raise ValueError(f"Non-numeric sheet index: {sheet_idx}")
        if not idx_float.is_integer():
            raise ValueError(f"Non-integer sheet index: {sheet_idx}")
        idx = int(idx_float)
    if idx < 1 or idx > len(wb_bdap.sheetnames):
        raise ValueError(f"Invalid BDAP sheet index: {idx}. Workbook has {len(wb_bdap.sheetnames)} sheets")
    return wb_bdap[wb_bdap.sheetnames[idx - 1]]


def _resolve_dynamic_column(ws_bdap, target_year: int, base_row: int, special_key: Optional[str], start_col: int = 1) -> Optional[int]:
    """Trova dinamicamente la colonna dell'anno e gestisce il fallback residui."""
    y_col = find_year_column_near_row(ws_bdap, target_year, base_row, start_col=start_col)
    if y_col is not None:
        return y_col

    if special_key and "residui" in special_key.lower():
        first_available_year = None
        start_row = max(1, base_row - 6)

        for r in range(start_row, base_row):
            for c in range(start_col, ws_bdap.max_column + 1):
                y = _coerce_year_value(ws_bdap.cell(row=r, column=c).value)
                if y and (first_available_year is None or y < first_available_year):
                    first_available_year = y

        if first_available_year and target_year == first_available_year - 1:
            for r in range(start_row, base_row):
                for c in range(start_col, ws_bdap.max_column + 1):
                    val = ws_bdap.cell(row=r, column=c).value
                    if isinstance(val, str) and "esercizi precedenti" in normalize_text(val):
                        return c
    return None


def _table_has_year_headers(ws_bdap, base_row: int, max_search_rows: int = 6) -> bool:
    """Controlla se ci sono colonne con intestazioni di anni valide vicino alla riga base."""
    start_row = max(1, base_row - max_search_rows)
    for row_idx in range(start_row, base_row):
        for col_idx in range(1, ws_bdap.max_column + 1):
            if _coerce_year_value(ws_bdap.cell(row=row_idx, column=col_idx).value) is not None:
                return True
    return False


def _is_manual_review_only_label(normalized_row_label: str) -> bool:
    """Verifica se l'etichetta richiede solo revisione manuale."""
    for token in MANUAL_REVIEW_ONLY_LABELS:
        if token in normalized_row_label:
            return True
    return False


def _resolve_special_source_entry(normalized_row_label: str, sources: Optional[dict[str, dict]] = None) -> tuple[Optional[str], Optional[dict]]:
    """Risolve la regola di sorgente speciale e restituisce la chiave e la mappatura della sorgente."""
    if not normalized_row_label:
        return None, None
    mapping_sources = sources or {}
    if normalized_row_label in mapping_sources:
        return normalized_row_label, mapping_sources[normalized_row_label]

    def _source_matches_row_context(key: str) -> bool:
        key_norm = normalize_text(key)
        return "fcde" not in key_norm or "fcde" in normalized_row_label

    # Quick check: if the normalized label contains 'titolo' followed by a roman
    # numeral (e.g. 'titoloiii'), try to map directly to a key that encodes the
    # spelled-out number (e.g. 'tre') to avoid fuzzy collisions.
    roman_to_it_quick = {'i': 'uno', 'ii': 'due', 'iii': 'tre', 'iv': 'quattro', 'v': 'cinque'}
    for r in sorted(roman_to_it_quick.keys(), key=len, reverse=True):
        if f"titolo{r}" in normalized_row_label:
            word = roman_to_it_quick[r]
            for key, source in mapping_sources.items():
                if not _source_matches_row_context(key):
                    continue
                if word in normalize_text(key) and 'titolo' in normalize_text(key):
                    return key, source

    # 1. Prova a trovare una chiave che sia contenuta nell'etichetta o che contenga l'etichetta (match parziale semplice)
    for key, source in mapping_sources.items():
        if not _source_matches_row_context(key):
            continue
        key_norm = normalize_text(key)
        if key_norm in normalized_row_label or normalized_row_label in key_norm:
            return key, source

    # 2. Fallback fuzzy matching con punteggio di similarità e sovrapposizione token, per gestire etichette simili.
    scores = []
    for key, source in mapping_sources.items():
        if not _source_matches_row_context(key):
            continue
        score = label_match_score(key, normalized_row_label)
        key_tokens = set(tokenize_text(key))
        label_tokens = set(tokenize_text(normalized_row_label))
        token_overlap = len(key_tokens & label_tokens)
        scores.append((score, token_overlap, key, source))

    if not scores:
        return None, None

    # Ordina prima per punteggio di similarità, poi per sovrapposizione token, entrambi in ordine decrescente
    scores.sort(key=lambda t: (t[0], t[1]), reverse=True)

    # Se ci sono più candidati con lo stesso punteggio e overlap, proviamo a risolvere il pareggio dando priorità a chi ha un numero 
    # romano che corrisponde a un token nel label. Questo aiuta a distinguere ad esempio "residui attivi titolo III" da "residui attivi titolo I" quando entrambi hanno una buona similarità.
    top_score = scores[0][0]
    top_overlap = scores[0][1]
    tied = [s for s in scores if s[0] == top_score and s[1] == top_overlap]
    chosen_from_tied = None
    if len(tied) > 1:
        # Detect roman numerals by substring in the normalized label (covers cases
        # like 'titoloiii' where tokenization produces a single long token).
        roman_to_it = {'i': 'uno', 'ii': 'due', 'iii': 'tre', 'iv': 'quattro', 'v': 'cinque'}
        # Prefer longest matches first (e.g. 'iii' before 'i')
        roman_in_label = sorted([r for r in roman_to_it.keys() if r in normalized_row_label], key=len, reverse=True)
        if roman_in_label:
            preferred = None
            for r in roman_in_label:
                word = roman_to_it[r]
                for score, overlap, key, source in tied:
                    key_tokens = set(tokenize_text(key))
                    if word in key_tokens:
                        preferred = (score, overlap, key, source)
                        break
                if preferred:
                    break
            if preferred:
                best_score, best_overlap, best_key, best_source = preferred
                if best_score >= 0.84:
                    chosen_from_tied = (best_key, best_source)

    # Se non abbiamo trovato un vincitore chiaro tra i candidati in parità, ma un numero romano
    # appare nell'etichetta, permette la selezione di un candidato il cui token numerico scritto
    # corrisponde al numero romano.
    # Also check for roman numerals as substrings in the normalized label
    roman_to_it = {'i': 'uno', 'ii': 'due', 'iii': 'tre', 'iv': 'quattro', 'v': 'cinque'}
    # Prefer longest matches first (e.g. 'iii' before 'i')
    roman_in_label = sorted([r for r in roman_to_it.keys() if r in normalized_row_label], key=len, reverse=True)
    if roman_in_label:
        best_score_overall = scores[0][0]
        for r in roman_in_label:
            word = roman_to_it[r]
            for score, overlap, key, source in scores:
                key_tokens = set(tokenize_text(key))
                if word in key_tokens and score >= (best_score_overall - 0.06):
                    if score >= 0.84:
                        return key, source
    # If the roman-fallback didn't return a match but we found a preferred
    # candidate among tied entries earlier, return it now.
    if chosen_from_tied is not None:
        return chosen_from_tied

    best_score, best_overlap, best_key, best_source = scores[0]
    if best_key is not None and best_score >= 0.84:
        return best_key, best_source
    return None, None


def _find_sheet_by_name_contains(wb_bdap, token: str):
    """Trova un worksheet il cui nome contiene il token fornito."""
    token_norm = normalize_text(token)
    for name in wb_bdap.sheetnames:
        if token_norm in normalize_text(name):
            return wb_bdap[name]
    return None


def _build_special_sources(ws_analysis) -> dict[str, dict]:
    """Costruisce le sorgenti utilizzando solo le mappature definite lato Python."""
    try:
        from bdap_app.support.default_mappings import TEMPLATE_SOURCES
    except Exception:
        TEMPLATE_SOURCES = {}

    if not isinstance(TEMPLATE_SOURCES, dict):
        return {}

    return {key: dict(value) for key, value in TEMPLATE_SOURCES.items()}


def _has_calculation_fallback(source_mapping: dict) -> bool:
    """Indica se la sorgente ha un fallback di calcolo da tentare prima di dichiararla irrisolta."""
    return any(key in source_mapping for key in ("compute_if_missing", "fallback_ratio_cells"))


def _adjust_cell_ref_to_row(cell_ref: str, target_row: int) -> str:
    """Adjust cell reference to use a different row number."""
    if not cell_ref:
        return cell_ref
    match = re.match(r"([A-Z]+)(\d+)", cell_ref.upper())
    if match:
        col = match.group(1)
        return f"{col}{target_row}"
    return cell_ref


def _resolve_scarto_formula_reference(formula: str) -> Optional[tuple[str, int, int]]:
    """Extract sheet, row offset and col offset from SCARTO/OFFSET formula pattern."""
    if not isinstance(formula, str):
        return None

    formula_up = formula.upper()
    if "SCARTO(" not in formula_up and "OFFSET(" not in formula_up:
        return None

    matches = re.findall(
        r"(?:_XLFN\.)?(?:SCARTO|OFFSET)\(\s*([^!]+)!\$?([A-Z]+)\$?(\d+)\s*[;,]\s*(-?\d+)\s*[;,]\s*(-?\d+)\s*\)",
        formula,
        flags=re.IGNORECASE,
    )
    if not matches:
        return None

    sheet_token, base_col, base_row, row_off, col_off = matches[-1]
    sheet_name = str(sheet_token).strip().strip("'").strip('"')
    try:
        base_col_idx = column_index_from_string(base_col.upper())
        target_row = int(base_row) + int(row_off)
        target_col = base_col_idx + int(col_off)
    except Exception:
        return None

    if target_row < 1 or target_col < 1:
        return None
    return sheet_name, target_row, target_col


def fill_accounting_analysis(
    analysis_path: Path,
    bdap_path: Optional[Path],
    output_path: Path,
    year: int,
    analysis_sheet: str = "analisi contabile",
    reference_year: Optional[int] = None,
    clear_when_missing: bool = False,
    keep_source_reference: bool = True,
    progress: Optional[Callable[[int, int], None]] = None,
):
    """Riempe la colonna dell'anno target nel foglio di analisi risolvendo i riferimenti BDAP."""
    wb_analysis = load_workbook(analysis_path) # Template di analisi contabile
    wb_bdap = None
    if not clear_when_missing:
        if bdap_path is None:
            raise ValueError("bdap_path is required unless clear_when_missing=True")
        wb_bdap = load_workbook(bdap_path, data_only=True)

    ws_analysis = find_sheet(wb_analysis, analysis_sheet) # Foglio di analisi contabile
    year_col = find_year_column(ws_analysis, year) # Colonna dell'anno da popolare nel foglio di analisi contabile
    if year_col is None:
        raise ValueError(f"Year column {year} not found in sheet '{analysis_sheet}'")

    reference_col = year_col
    if reference_year is not None:
        ref_year_col = find_year_column(ws_analysis, reference_year)
        if ref_year_col is not None:
            reference_col = ref_year_col
    if reference_col == year_col:
        fallback_reference_col = find_reference_column(ws_analysis)
        if fallback_reference_col is not None:
            reference_col = fallback_reference_col

    updated_rows = 0
    unresolved_references = 0
    unresolved_by_reason = {}
    resolved_special_numeric_values = {}
    special_sources = _build_special_sources(ws_analysis)
    wb_questionario = None
    wb_questionario_formula = None
    wb_indicatori = None
    questionario_workbook_cache: dict[str, tuple[object, object]] = {}

    def mark_unresolved(target_cell, reason: str) -> None:
        nonlocal unresolved_references
        target_cell.value = None
        target_cell.comment = None
        unresolved_references += 1
        unresolved_by_reason[reason] = unresolved_by_reason.get(reason, 0) + 1

    max_row = ws_analysis.max_row or 0
    progress_interval = max(1, max_row // 100) if max_row else 1
    if progress is not None:
        progress(0, max_row)

    for row_idx in range(1, max_row + 1):
        if progress is not None and (
            row_idx == 1 or row_idx == max_row or row_idx % progress_interval == 0
        ):
            progress(row_idx, max_row)

        target_cell = ws_analysis.cell(row=row_idx, column=year_col)
        if isinstance(target_cell, MergedCell):
            continue
        
        reference_text = ws_analysis.cell(row=row_idx, column=reference_col).value
        row_label = extract_row_label(ws_analysis, row_idx)
        normalized_row_label = normalize_text(row_label)

        parsed = parse_reference(reference_text)
        if not parsed:
            parsed = parse_reference(target_cell.value)

        if not normalized_row_label and not parsed:
            continue
        if row_idx == 1 and normalized_row_label in {"voce", "anno", "descrizione", "riferimento", "label"}:
            continue

        special_key, special_source = _resolve_special_source_entry(normalized_row_label, special_sources)
        use_special_source = special_source is not None
        source_workbook = normalize_text(special_source.get("source_workbook", "rendiconto")) if use_special_source else "rendiconto"

        # Flusso da diagramma: senza mapping deterministico e senza riferimento BDAP esplicito
        # la riga non viene compilata automaticamente.
        if not use_special_source and parsed is None:
            continue

        # Flusso da diagramma: le righe da relazione non vengono risolte automaticamente.
        # La cella resta vuota e il caso viene contato tra gli unresolved.
        if use_special_source and source_workbook == "relazione":
            process_relazione_row(
                target_cell,
                bdap_path,
                year,
                special_key,
                special_source,
                keep_source_reference=keep_source_reference,
            )
            mark_unresolved(target_cell, f"source_workbook relazione: {special_key or row_label}")
            continue

        if _is_manual_review_only_label(normalized_row_label):
            continue

        sheet_token = parsed[0] if parsed else None

        try:
            # Se siamo in modalità clear_when_missing, non tentiamo di risolvere nulla e cancelliamo direttamente il valore
            if clear_when_missing: 
                target_cell.value = None
                updated_rows += 1
                continue

            assert wb_bdap is not None 

            if use_special_source:
                actual_cell_ref = None
                ws_bdap_src = None
                source_descr_override = None
                source_file_name = bdap_path.name if bdap_path is not None else "n/d"

                source_wb = wb_bdap

                if source_workbook in DELEGATED_SOURCE_HANDLERS:
                    handler = DELEGATED_SOURCE_HANDLERS[source_workbook]
                    handler_kwargs = {}
                    if source_workbook == "questionario":
                        handler_kwargs["workbook_cache"] = questionario_workbook_cache
                    handled, delta_rows, wb_x, wb_x_formula, _path_cache, _ = handler(
                        target_cell,
                        bdap_path,
                        year,
                        special_key,
                        special_source,
                        keep_source_reference=keep_source_reference,
                        **handler_kwargs,
                    )

                    # store discovered workbooks for cleanup
                    if source_workbook == "questionario" and wb_x is not None:
                        wb_questionario = wb_x
                        wb_questionario_formula = wb_x_formula
                    if source_workbook == "indicatori" and wb_x is not None:
                        wb_indicatori = wb_x

                    if handled:
                        updated_rows += delta_rows
                        continue
                    raise KeyError(f"Missing {source_workbook} workbook")

                if special_key == "fcdeinrapportoaresiduiattivi":
                    fcde_value, fcde_cell_ref = read_fcde_percentage_from_workbook(wb_analysis, year)
                    if fcde_value is None and bdap_path is not None:
                        fcde_value, fcde_cell_ref = populate_fcde_sheet_in_workbook(
                            wb_analysis,
                            bdap_path,
                            year,
                            keep_source_reference=keep_source_reference,
                        )
                    if fcde_value is None and wb_bdap is not None:
                        fcde_value, fcde_cell_ref, _ = _resolve_fcde_ratio_from_source(wb_bdap)
                    if fcde_value is not None:
                        formatted_value = format_value_italian(
                            fcde_value,
                            percent=bool(special_source.get("is_percentage", False)),
                        )
                        target_cell.value = formatted_value
                        if keep_source_reference and bdap_path is not None:
                            source_file_name = bdap_path.name
                            source_descr = f"foglio FCDE, cella {fcde_cell_ref or 'FCDE!I6'}"
                            target_cell.comment = Comment(f"Fonte BDAP: {source_file_name} | {source_descr}", "BDAP Automation")
                        updated_rows += 1
                        continue
                    if special_source.get("optional", False):
                        mark_unresolved(target_cell, f"No value found for {special_key or row_label}")
                        continue

                assert source_wb is not None

                if "sheet_name_contains" in special_source:
                    ws_bdap_src = _find_sheet_by_name_contains(source_wb, special_source["sheet_name_contains"])
                    if ws_bdap_src is None and "sheet_idx" in special_source:
                        ws_bdap_src = resolve_bdap_sheet(source_wb, special_source["sheet_idx"])
                    if ws_bdap_src is None and not special_source.get("optional", False):
                        raise KeyError(f"Missing sheet matching '{special_source['sheet_name_contains']}'")
                elif "sheet_idx" in special_source:
                    ws_bdap_src = resolve_bdap_sheet(source_wb, special_source["sheet_idx"])

                if ws_bdap_src is None:
                    if special_source.get("optional", False):
                        mark_unresolved(target_cell, f"Missing sheet for {special_key or row_label}")
                        continue
                    continue

                value = None
                display_refs_handled = False

                if "first_non_zero" in special_source:
                    row_tol = max(0, _coerce_int_or_default(special_source.get("row_tolerance", 2), 2))

                    for option in special_source["first_non_zero"]:
                        ws_opt = resolve_bdap_sheet(source_wb, option["sheet_idx"])
                        candidate, candidate_ref = find_bdap_value_with_fallback(ws_opt, option["cell_ref"], max_offset=row_tol)

                        if coerce_numeric(candidate) not in (None, 0):
                            value = candidate
                            ws_bdap_src = ws_opt
                            actual_cell_ref = candidate_ref
                            break
                            
                    if value is None:
                        option = special_source["first_non_zero"][0]
                        ws_bdap_src = resolve_bdap_sheet(source_wb, option["sheet_idx"])
                        value, actual_cell_ref = find_bdap_value_with_fallback(ws_bdap_src, option["cell_ref"], max_offset=row_tol)
                
                elif "sum_cells" in special_source:
                    total = Decimal("0")
                    found_any = False
                    last_ref = None

                    row_tol = max(0, _coerce_int_or_default(special_source.get("row_tolerance", 1), 1))
                    search_prefix = special_source.get("search_key_prefix", None)
                    label_check_cells = special_source.get("label_check_cells", None)

                    expected_label_cfg = special_source.get("expected_label")

                    # Se sum_cells contiene solo la la colonna:
                    sum_cells_cfg = list(special_source.get("sum_cells", []))
                    col_only = [c for c in sum_cells_cfg if isinstance(c, str) and re.match(r'^[A-Za-z]+$', c.strip())]
                    if col_only:
                        for col_letter in col_only:
                            col_idx = column_index_from_string(col_letter.upper())
                            # scan rows for label occurrence
                            for r in range(1, ws_bdap_src.max_row + 1):
                                row_text = ""
                                max_check_cols = min(6, ws_bdap_src.max_column)
                                for c_idx in range(1, max_check_cols + 1):
                                    val = ws_bdap_src.cell(row=r, column=c_idx).value
                                    if isinstance(val, str):
                                        row_text += val + " "

                                match_found = False
                                if isinstance(expected_label_cfg, (list, tuple)):
                                    for lbl in expected_label_cfg:
                                        if lbl and normalize_text(str(lbl)) in normalize_text(row_text):
                                            match_found = True
                                            break
                                elif expected_label_cfg:
                                    if normalize_text(str(expected_label_cfg)) in normalize_text(row_text):
                                        match_found = True

                                if not match_found:
                                    continue

                                candidate = ws_bdap_src.cell(row=r, column=col_idx).value
                                if (num_candidate := coerce_numeric(candidate)) is not None:
                                    total += num_candidate
                                    found_any = True
                                    last_ref = f"{col_letter.upper()}{r}"

                        value = total if found_any else None
                        actual_cell_ref = last_ref
                    else:
                        # Lista o tupla
                        if isinstance(expected_label_cfg, (list, tuple)):
                            labels = list(expected_label_cfg)
                            sum_cells = sum_cells_cfg
                            for idx, lbl in enumerate(labels):
                                lbl_str = str(lbl).strip()
                                target_row_for_label = find_row_by_expected_label(
                                    ws_bdap_src, lbl_str, row_tolerance=row_tol, search_key_prefix=search_prefix, label_check_cells=label_check_cells
                                )
                                if target_row_for_label is None or idx >= len(sum_cells):
                                    continue

                                cell_ref = sum_cells[idx]

                                actual_cell_ref = _adjust_cell_ref_to_row(cell_ref, target_row_for_label) if target_row_for_label else cell_ref
                                candidate, candidate_ref = find_bdap_value_with_fallback(ws_bdap_src, actual_cell_ref, max_offset=2)
                                if (num_candidate := coerce_numeric(candidate)) is not None:
                                    total += num_candidate
                                    found_any = True
                                    last_ref = candidate_ref or actual_cell_ref

                        else:
                            target_row = None
                            if expected_label_cfg:
                                expected_label = str(expected_label_cfg).strip()
                                target_row = find_row_by_expected_label(
                                    ws_bdap_src, expected_label, row_tolerance=row_tol, search_key_prefix=search_prefix, label_check_cells=label_check_cells
                                )

                            for cell_ref in sum_cells_cfg:
                                actual_cell_ref = _adjust_cell_ref_to_row(cell_ref, target_row) if target_row else cell_ref
                                candidate, candidate_ref = find_bdap_value_with_fallback(ws_bdap_src, actual_cell_ref, max_offset=2)
                                if (num_candidate := coerce_numeric(candidate)) is not None:
                                    total += num_candidate
                                    found_any = True
                                    last_ref = candidate_ref or actual_cell_ref

                        value = total if found_any else None
                        actual_cell_ref = last_ref

                elif special_key == "fcdeinrapportoaresiduiattivi":
                    # Il valore viene già risolto dal foglio FCDE compilato prima dell'analisi.
                    pass
                
                elif "display_refs" in special_source:
                    display_refs = special_source["display_refs"]
                    use_static_refs = parse_bool_flag(special_source.get("static_column"), default=False)
                    row_tol = max(0, _coerce_int_or_default(special_source.get("row_tolerance", 2), 2))
                    lines = []
                    last_refs = []
                    logger.debug(
                        f"Processing display_refs for special_key={special_key} display_refs={display_refs} on sheet {ws_bdap_src.title if ws_bdap_src is not None else 'n/d'}"
                    )
                    for title, cell_ref in display_refs:
                        try:
                            base_row, base_col = coordinate_to_tuple(cell_ref)
                        except Exception:
                            base_row, base_col = None, None

                        candidate = None
                        candidate_ref = None
                        year_aborted = False

                        if base_row is not None:
                            target_row = base_row
                            found_title_row = find_nearby_title_row(
                                ws_bdap_src,
                                str(title),
                                base_row,
                                base_col,
                                row_tolerance=row_tol,
                            )
                            if found_title_row is not None:
                                target_row = found_title_row

                            # If the mapping provides explicit year_cell_ref (possibly as a list),
                            # prefer that to resolve the year column.
                            y_col = None
                            if not use_static_refs and "year_cell_ref" in special_source:
                                refs = special_source["year_cell_ref"]
                                if isinstance(refs, (list, tuple)):
                                    for yref in refs:
                                        try:
                                            y_row, y_col_idx = coordinate_to_tuple(yref)
                                            val = ws_bdap_src.cell(row=y_row, column=y_col_idx).value
                                            if _coerce_year_value(val) == year:
                                                y_col = y_col_idx
                                                break
                                        except Exception:
                                            continue
                                else:
                                    try:
                                        y_row, y_col_idx = coordinate_to_tuple(refs)
                                        val = ws_bdap_src.cell(row=y_row, column=y_col_idx).value
                                        if _coerce_year_value(val) == year:
                                            y_col = y_col_idx
                                    except Exception:
                                        y_col = None
                            if not use_static_refs and y_col is None:
                                y_col = _resolve_dynamic_column(ws_bdap_src, year, target_row, special_key)
                            if y_col is not None:
                                year_cell = ws_bdap_src.cell(row=target_row, column=y_col)
                                candidate = year_cell.value
                                candidate_ref = year_cell.coordinate
                            else:
                                # If an explicit year column could not be resolved, try a
                                # conservative fallback: read the static base `cell_ref`
                                # (no offset). This helps when older questionario files
                                # store values in fixed columns without a clear year header
                                # and avoids leaving the target empty for years like 2021.
                                try:
                                    actual_ref = f"{get_column_letter(base_col)}{target_row}" if base_col else cell_ref
                                    candidate, candidate_ref = find_bdap_value_with_fallback(ws_bdap_src, actual_ref, max_offset=0)
                                except Exception:
                                    # If fallback also fails, mark as aborted only when
                                    # the table actually contains year headers (to avoid
                                    # masking real missing-year situations).
                                    if "year_cell_ref" in special_source or _table_has_year_headers(ws_bdap_src, base_row):
                                        year_aborted = True
                                    else:
                                        candidate = None
                                        candidate_ref = None

                        logger.debug(
                            f"display_refs: title={title}, cell_ref={cell_ref}, candidate={candidate}, candidate_ref={candidate_ref}, year_aborted={year_aborted}"
                        )
                        # Always include the title line; missing or aborted -> show 0
                        if year_aborted:
                            num_val = Decimal('0')
                            last_refs.append(cell_ref)
                        else:
                            coerced = coerce_numeric(candidate)
                            if coerced is None:
                                num_val = Decimal('0')
                                last_refs.append(cell_ref)
                            else:
                                num_val = coerced
                                last_refs.append(candidate_ref or cell_ref)

                        formatted = format_value_italian(num_val)
                        lines.append(f"{title}: {formatted}")

                    value = "\n".join(lines)
                    actual_cell_ref = ";".join(str(r) for r in last_refs) if last_refs else "DISPLAY_REFS"
                    source_descr_override = f"foglio {ws_bdap_src.title} | " + " | ".join(f"{title}: cella {cell_ref}" for title, cell_ref in display_refs)
                    display_refs_handled = True
                
                elif "ratio_cells" in special_source:
                    num_ref, den_ref = special_source["ratio_cells"]
                    numerator, num_actual = find_bdap_value_with_fallback(ws_bdap_src, num_ref, max_offset=0)
                    denominator, den_actual = find_bdap_value_with_fallback(ws_bdap_src, den_ref, max_offset=0)
                    if (num_coerce := coerce_numeric(numerator)) and (den_coerce := coerce_numeric(denominator)) and den_coerce != 0:
                        value = num_coerce / den_coerce
                        if parse_bool_flag(special_source.get("ratio_percent", True)):
                            value *= Decimal("100")
                        actual_cell_ref = "CALCULATED_RATIO"
                        source_descr_override = f"foglio {ws_bdap_src.title}, formula {num_actual or num_ref}/{den_actual or den_ref}"
                        if parse_bool_flag(special_source.get("ratio_percent", True)):
                            source_descr_override += " * 100"
                
                elif "cell_refs" in special_source and "cell_ref" not in special_source:
                    # Caso: solo cell_refs (senza cell_ref), come anticipazionetesoreria
                    row_tol = max(0, _coerce_int_or_default(special_source.get("row_tolerance", 1), 1))
                    expected_label = str(special_source.get("expected_label", "")).strip()
                    label_check_cells = special_source.get("label_check_cells")
                    cell_refs = special_source.get("cell_refs", [])
                    
                    if expected_label and label_check_cells:
                        label_check_cells_list = label_check_cells if isinstance(label_check_cells, list) else [label_check_cells]
                        label_found_idx = None
                        target_norm = normalize_text(expected_label)
                        
                        # Cerca quale label_check_cell contiene l'etichetta
                        for idx, check_cell_ref in enumerate(label_check_cells_list):
                            try:
                                # Handle range notation like 'A25:C25'
                                if ':' in str(check_cell_ref):
                                    parts = str(check_cell_ref).split(':')
                                    check_cell_ref = parts[0]
                                
                                check_row, check_col = coordinate_to_tuple(check_cell_ref)
                                
                                for r_offset in range(-row_tol, row_tol + 1):
                                    curr_row = check_row + r_offset
                                    if curr_row < 1:
                                        continue
                                    
                                    cell_val = ws_bdap_src.cell(row=curr_row, column=check_col).value
                                    if isinstance(cell_val, str) and target_norm in normalize_text(cell_val):
                                        label_found_idx = idx
                                        break
                                
                                if label_found_idx is not None:
                                    break
                            except Exception:
                                continue
                        
                        # Se trovato, leggi da cell_refs[indice]
                        if label_found_idx is not None and label_found_idx < len(cell_refs):
                            value_cell_ref = cell_refs[label_found_idx]
                            value, actual_cell_ref = find_bdap_value_with_fallback(
                                ws_bdap_src,
                                value_cell_ref,
                                max_offset=row_tol
                            )
                        else:
                            # Se non trovato, fallback: prova prima cella in cell_refs
                            if cell_refs:
                                value, actual_cell_ref = find_bdap_value_with_fallback(
                                    ws_bdap_src,
                                    cell_refs[0],
                                    max_offset=row_tol
                                )
                            else:
                                value = None
                    else:
                        value = None
                
                elif "cell_ref" in special_source:
                    row_tol = max(0, _coerce_int_or_default(special_source.get("row_tolerance", 1), 1))
                    expected_labels = text_options(special_source.get("expected_label"))
                    label_check_cells = special_source.get("label_check_cells")
                    strict_expected_label = bool(special_source.get("strict_expected_label", False))
                    search_key_prefix = special_source.get("search_key_prefix", None)
                    
                    actual_base_ref = special_source["cell_ref"]
                    fallback_refs = special_source.get("fallback_cell_refs", [])
                    if isinstance(fallback_refs, str):
                        fallback_refs = [fallback_refs]
                    candidate_base_refs = [actual_base_ref] + [ref for ref in fallback_refs if ref]
                    is_static = special_source.get("static_column", False)
                    check_year_in_label = special_source.get("check_year_in_label", False)
                    
                    if (not is_static) or "year_cell_ref" in special_source:
                        try:
                            b_row, _ = coordinate_to_tuple(actual_base_ref)
                            y_col = _resolve_dynamic_column(ws_bdap_src, year, b_row, special_key)
                            
                            if y_col is not None:
                                actual_base_ref = f"{get_column_letter(y_col)}{b_row}"
                            else:
                                if "year_cell_ref" in special_source or _table_has_year_headers(ws_bdap_src, b_row):
                                    if not _has_calculation_fallback(special_source):
                                        mark_unresolved(target_cell, f"Year column {year} not found for {special_key or row_label}")
                                        continue
                        except Exception:
                            pass

                    if expected_labels:
                        if check_year_in_label:
                            # Logica esatta richiesta dall'utente: cerca l'etichetta E controlla che contenga l'anno
                            found_match = False
                            for base_ref in candidate_base_refs:
                                match = re.match(r"([A-Z]+)(\d+)", str(base_ref).upper())
                                if not match:
                                    continue
                                b_col_letter = match.group(1)
                                b_row = int(match.group(2))
                                b_col_idx = column_index_from_string(b_col_letter)

                                for r_offset in range(-row_tol, row_tol + 1):
                                    curr_row = b_row + r_offset
                                    if curr_row < 1:
                                        continue

                                    row_text = ""
                                    # Leggiamo il testo presente nelle celle delle etichette (es. A, B, C)
                                    for c_idx in range(1, 6):
                                        val = ws_bdap_src.cell(row=curr_row, column=c_idx).value
                                        if isinstance(val, str):
                                            row_text += val.lower() + " "

                                    # Controlla se la frase base esiste E se l'anno (es. "2024") e' nella stessa frase
                                    if any(normalize_text(label) in normalize_text(row_text) for label in expected_labels) and str(year) in row_text:
                                        val_cell = ws_bdap_src.cell(row=curr_row, column=b_col_idx)
                                        value = val_cell.value
                                        actual_cell_ref = val_cell.coordinate
                                        actual_base_ref = base_ref
                                        found_match = True
                                        break
                                if found_match:
                                    break

                            if not found_match:
                                value = None
                        else:
                            for base_ref in candidate_base_refs:
                                for expected_label in expected_labels:
                                    value, actual_cell_ref = find_bdap_value_by_expected_label_near_cell(
                                        ws_bdap_src,
                                        base_ref,
                                        expected_label=expected_label,
                                        max_row_offset=row_tol,
                                        label_check_cells=label_check_cells,
                                        strict_expected_label=strict_expected_label,
                                        search_key_prefix=search_key_prefix,
                                    )
                                    if value is not None:
                                        actual_base_ref = base_ref
                                        break
                                if value is not None:
                                    break
                    
                    if expected_labels and value is None:
                        # La label attesa valida la riga quando possibile, ma non deve
                        # impedire l'uso della cella deterministica del mapping.
                        pass

                    if value is None and not special_source.get("only_fill_if_found", False):
                        for base_ref in candidate_base_refs:
                            value, actual_cell_ref = find_bdap_value_with_fallback(ws_bdap_src, base_ref, max_offset=row_tol)
                            if value is not None:
                                actual_base_ref = base_ref
                                break
                
                else:
                    value, actual_cell_ref = find_bdap_value_by_label(ws_bdap_src, special_source["label"])

                if value is None and "fallback_ratio_cells" in special_source:
                    ratio_ws = ws_bdap_src
                    if "fallback_ratio_sheet_name_contains" in special_source:
                        ratio_ws = _find_sheet_by_name_contains(source_wb, special_source["fallback_ratio_sheet_name_contains"])
                    elif "fallback_ratio_sheet_idx" in special_source:
                        ratio_ws = resolve_bdap_sheet(source_wb, special_source["fallback_ratio_sheet_idx"])
                    if ratio_ws is not None:
                        num_ref, den_ref = special_source["fallback_ratio_cells"]
                        numerator, num_actual = find_bdap_value_with_fallback(ratio_ws, num_ref, max_offset=0)
                        denominator, den_actual = find_bdap_value_with_fallback(ratio_ws, den_ref, max_offset=0)
                        if (num := coerce_numeric(numerator)) and (den := coerce_numeric(denominator)) and den != 0:
                            value = num / den
                            if parse_bool_flag(special_source.get("fallback_ratio_percent", True)):
                                value *= Decimal("100")
                            actual_cell_ref = "CALCULATED_FALLBACK_RATIO"
                            source_descr_override = f"foglio {ratio_ws.title}, fallback formula {num_actual or num_ref}/{den_actual or den_ref}"
                            if parse_bool_flag(special_source.get("fallback_ratio_percent", True)):
                                source_descr_override += " * 100"

                if value is None and "compute_if_missing" in special_source:
                    compute_rule = special_source["compute_if_missing"]
                    base_key = compute_rule.get("base")
                    subtract_keys = list(compute_rule.get("subtract", []))
                    if base_key in resolved_special_numeric_values and all(key in resolved_special_numeric_values for key in subtract_keys):
                        computed = resolved_special_numeric_values[base_key]
                        for key in subtract_keys:
                            computed -= resolved_special_numeric_values[key]
                        value = computed
                        actual_cell_ref = "CALCULATED"

                if value is None and "default_if_missing" in special_source:
                    value = special_source["default_if_missing"]
                    actual_cell_ref = "DEFAULT_IF_MISSING"

                if value is None:
                    fallback_label = first_text_option(special_source.get("label")) or first_text_option(special_source.get("expected_label"))
                    if fallback_label and not special_source.get("only_fill_if_found", False):
                        value, actual_cell_ref = find_bdap_value_by_label(ws_bdap_src, fallback_label)

                if value is None:
                    if "cell_ref" in special_source:
                        mark_unresolved(target_cell, f"No value found in cell {special_source['cell_ref']}")
                    else:
                        mark_unresolved(target_cell, f"No value found for {special_key or row_label}")
                    continue

                # Respect percentage flag in the source mapping
                percent_flag = bool(special_source.get("is_percentage", False))
                formatted_value = format_value_italian(value, percent=percent_flag)
                target_cell.value = formatted_value
                # If we wrote a multiline display_refs value, enable wrapText so it displays
                try:
                    if display_refs_handled:
                        target_cell.alignment = Alignment(wrapText=True)
                    elif isinstance(formatted_value, str) and "\n" in formatted_value:
                        target_cell.alignment = Alignment(wrapText=True)
                except Exception:
                    pass
                if (numeric_value := coerce_numeric(value)) is not None and special_key is not None:
                    resolved_special_numeric_values[special_key] = numeric_value
                    if special_key == "prospettoevoluzionerisultatoammne":
                        resolved_special_numeric_values["risultatoamministrazionea"] = numeric_value
                    if special_key == "risultatoamministrazionea":
                        resolved_special_numeric_values["prospettoevoluzionerisultatoammne"] = numeric_value

                if keep_source_reference and bdap_path is not None:
                    if source_descr_override is not None:
                        source_descr = source_descr_override
                    elif actual_cell_ref == "CALCULATED":
                        missing_ref = str(special_source.get("cell_ref", "sorgente prevista"))
                        source_descr = f"valore non trovato in {missing_ref} | calcolata come A-B-C-D"
                    elif actual_cell_ref == "DEFAULT_IF_MISSING":
                        source_descr = "voce non trovata nella sorgente | valore predefinito 0"
                    else:
                        source_descr = format_source_location(ws_bdap_src.title, actual_cell_ref)
                    if "label" in special_source:
                        source_descr += f" | voce {special_source['label']}"
                    target_cell.comment = Comment(f"Fonte BDAP: {source_file_name} | {source_descr}", "BDAP Automation")

                updated_rows += 1
                continue

            value = None
            actual_source = None

            if parsed is not None:
                ws_bdap_ref = resolve_bdap_sheet(wb_bdap, sheet_token)
                bdap_cell = parsed[1]
                value, actual_cell_ref = find_bdap_value_with_fallback(ws_bdap_ref, bdap_cell, max_offset=2)
                if value is not None:
                    actual_source = format_source_location(ws_bdap_ref.title, actual_cell_ref)

            if value is None:
                raise KeyError(f"No value found for BDAP reference {reference_text or target_cell.coordinate}")

            formatted_value = format_value_italian(value)
            target_cell.value = formatted_value
            
            if keep_source_reference and bdap_path is not None:
                comment_text = f"Fonte BDAP: {bdap_path.name}"
                if actual_source:
                    comment_text += f" | {actual_source}"
                elif sheet_token:
                    comment_text += f" | foglio {sheet_token}"
                target_cell.comment = Comment(comment_text, "BDAP Automation")
            
            updated_rows += 1

        except (ValueError, TypeError, KeyError, IndexError, AttributeError) as exc:
            reason = f"{type(exc).__name__}: {exc}"
            mark_unresolved(target_cell, reason)
            logger.warning("Unresolved BDAP reference at row %s (cell=%s, reference=%r): %s", row_idx, target_cell.coordinate, target_cell.value, reason)

    if unresolved_by_reason:
        logger.info("Unresolved BDAP reference summary:")
        for reason, count in sorted(unresolved_by_reason.items(), key=lambda item: item[1], reverse=True):
            logger.info("- %s -> %s", reason, count)

    wb_analysis.save(output_path)
    closed_workbook_ids = set()
    for wb_data, wb_formula in questionario_workbook_cache.values():
        for workbook in (wb_data, wb_formula):
            if workbook is not None and id(workbook) not in closed_workbook_ids:
                workbook.close()
                closed_workbook_ids.add(id(workbook))
    if wb_questionario is not None and id(wb_questionario) not in closed_workbook_ids:
        wb_questionario.close()
        closed_workbook_ids.add(id(wb_questionario))
    if wb_questionario_formula is not None and id(wb_questionario_formula) not in closed_workbook_ids:
        wb_questionario_formula.close()
        closed_workbook_ids.add(id(wb_questionario_formula))
    if wb_indicatori is not None:
        wb_indicatori.close()
    if wb_bdap is not None:
        wb_bdap.close()
    wb_analysis.close()
    return updated_rows, unresolved_references
