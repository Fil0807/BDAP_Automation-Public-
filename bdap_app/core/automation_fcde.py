"""Compilazione del foglio FCDE prima dell'analisi contabile."""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Callable, Optional

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

from .automation_search import find_year_column, resolve_sheet_by_contains_or_idx
from ..support.default_mappings import TEMPLATE_SOURCES
from ..support.text_utils import _coerce_int_or_default, text_options
from ..support.value_formatter import coerce_numeric
from ..support.value_resolver import (
    find_bdap_value_by_expected_label_near_cell,
    find_bdap_value_in_workbook_by_label,
    find_bdap_value_with_fallback,
)

FCDE_SHEET_NAME = "FCDE"
FCDE_ROW_IMPORTO = 2
FCDE_ROW_RESIDUI_TITOLO_I = 3
FCDE_ROW_RESIDUI_TITOLO_III = 4
FCDE_ROW_TOTALE = 5
FCDE_ROW_PERCENTUALE = 6
FCDE_ROW_METODO = 7

def _resolve_fcde_importo(source_wb) -> tuple[Optional[Decimal], Optional[str], Optional[str]]:
    mapping = dict(TEMPLATE_SOURCES.get("dicuifcde", {}))
    ws_source = resolve_sheet_by_contains_or_idx(source_wb, mapping, "sheet_name_contains", "sheet_idx")
    if ws_source is None:
        return None, None, None

    row_tol = max(0, _coerce_int_or_default(mapping.get("row_tolerance", 2), 2))
    cell_ref = str(mapping.get("cell_ref", "B28"))
    value, actual_ref = find_bdap_value_with_fallback(ws_source, cell_ref, max_offset=row_tol)
    numeric = coerce_numeric(value)
    if numeric is not None:
        return numeric, actual_ref, ws_source.title

    expected_label = mapping.get("expected_label")
    if expected_label:
        value, actual_ref = find_bdap_value_by_expected_label_near_cell(
            ws_source,
            cell_ref,
            expected_label=str(expected_label),
            max_row_offset=row_tol,
            label_check_cells=mapping.get("label_check_cells"),
            strict_expected_label=bool(mapping.get("strict_expected_label", False)),
        )
        numeric = coerce_numeric(value)
        if numeric is not None:
            return numeric, actual_ref, ws_source.title

    value, actual_ref = find_bdap_value_in_workbook_by_label(source_wb, str(expected_label or ""))
    numeric = coerce_numeric(value)
    if numeric is not None:
        return numeric, actual_ref, None
    return None, None, ws_source.title


def _resolve_fcde_residui_rendiconto(
    source_wb,
    special_key: str,
) -> tuple[Optional[Decimal], Optional[str], Optional[str]]:
    mapping = dict(TEMPLATE_SOURCES.get(special_key, {}))
    if not mapping:
        return None, None, None

    ws_source = resolve_sheet_by_contains_or_idx(source_wb, mapping, "sheet_name_contains", "sheet_idx")
    if ws_source is None:
        return None, None, None

    row_tol = max(0, _coerce_int_or_default(mapping.get("row_tolerance", 2), 2))
    cell_ref = str(mapping.get("cell_ref", ""))
    if not cell_ref:
        return None, None, ws_source.title

    expected_labels = text_options(mapping.get("expected_label"))
    for expected_label in expected_labels:
        value, actual_ref = find_bdap_value_by_expected_label_near_cell(
            ws_source,
            cell_ref,
            expected_label=expected_label,
            max_row_offset=row_tol,
            label_check_cells=mapping.get("label_check_cells"),
            strict_expected_label=bool(mapping.get("strict_expected_label", False)),
        )
        numeric = coerce_numeric(value)
        if numeric is not None:
            return numeric, actual_ref, ws_source.title

    value, actual_ref = find_bdap_value_with_fallback(ws_source, cell_ref, max_offset=row_tol)
    numeric = coerce_numeric(value)
    if numeric is not None:
        return numeric, actual_ref, ws_source.title

    for expected_label in expected_labels:
        value, source_descr = find_bdap_value_in_workbook_by_label(source_wb, expected_label)
        numeric = coerce_numeric(value)
        if numeric is not None:
            return numeric, source_descr, None

    return None, None, ws_source.title


def _resolve_fcde_ratio_from_source(source_wb) -> tuple[Optional[Decimal], Optional[str], Optional[str]]:
    mapping = dict(TEMPLATE_SOURCES.get("fcdeinrapportoaresiduiattivi", {}))
    ws_source = resolve_sheet_by_contains_or_idx(source_wb, mapping, "sheet_name_contains", "sheet_idx")
    if ws_source is not None:
        cell_ref = str(mapping.get("cell_ref", "I24"))
        row_tol = max(0, _coerce_int_or_default(mapping.get("row_tolerance", 2), 2))
        value, actual_ref = find_bdap_value_with_fallback(ws_source, cell_ref, max_offset=row_tol)
        numeric = coerce_numeric(value)
        if numeric is not None:
            return numeric, actual_ref, ws_source.title

        expected_label = mapping.get("expected_label") or "FCDE IN RAPPORTO A RESIDUI ATTIVI"
        value, actual_ref = find_bdap_value_by_expected_label_near_cell(
            ws_source,
            cell_ref,
            expected_label=str(expected_label),
            max_row_offset=row_tol,
            label_check_cells=mapping.get("label_check_cells"),
            strict_expected_label=bool(mapping.get("strict_expected_label", False)),
        )
        numeric = coerce_numeric(value)
        if numeric is not None:
            return numeric, actual_ref, ws_source.title

    value, source_descr = find_bdap_value_in_workbook_by_label(source_wb, "FCDE IN RAPPORTO A RESIDUI ATTIVI")
    numeric = coerce_numeric(value)
    if numeric is not None:
        return numeric, source_descr, None

    value, source_descr = find_bdap_value_in_workbook_by_label(
        source_wb,
        "% rispetto a residui attivi conservati (minimo 30%)",
    )
    numeric = coerce_numeric(value)
    if numeric is not None:
        return numeric, source_descr, None

    return None, None, None


def _read_fcde_percentage_from_sheet(ws_fcde, year: int) -> tuple[Optional[Decimal], Optional[str]]:
    year_col = find_year_column(ws_fcde, year, max_header_rows=3)
    if year_col is None:
        return None, None

    cell = ws_fcde.cell(row=FCDE_ROW_PERCENTUALE, column=year_col)
    numeric = coerce_numeric(cell.value)
    if numeric is not None:
        return numeric, cell.coordinate

    importo = coerce_numeric(ws_fcde.cell(row=FCDE_ROW_IMPORTO, column=year_col).value)
    totale = coerce_numeric(ws_fcde.cell(row=FCDE_ROW_TOTALE, column=year_col).value)
    if importo is None or totale in (None, Decimal("0")):
        return None, cell.coordinate
    return (importo * Decimal("100")) / totale, cell.coordinate


def _set_comment(target_cell, source_file_name: str, source_descr: str, keep_source_reference: bool) -> None:
    if not keep_source_reference:
        return
    target_cell.comment = Comment(f"Fonte BDAP: {source_file_name} | {source_descr}", "BDAP Automation")


def populate_fcde_sheet_in_workbook(
    wb_analysis,
    bdap_path: Optional[Path],
    year: int,
    keep_source_reference: bool = True,
    log: Optional[Callable[[str], None]] = None,
) -> tuple[Optional[Decimal], Optional[str]]:
    """Compila il foglio FCDE nel workbook già caricato."""
    if bdap_path is None:
        return None, None

    if FCDE_SHEET_NAME not in wb_analysis.sheetnames:
        return None, None

    ws_fcde = wb_analysis[FCDE_SHEET_NAME]
    year_col = find_year_column(ws_fcde, year, max_header_rows=3)
    if year_col is None:
        raise ValueError(f"Year column {year} not found in sheet '{FCDE_SHEET_NAME}'")

    source_file_name = bdap_path.name
    source_wb = load_workbook(bdap_path, data_only=True)
    try:
        importo_fcde, importo_ref, importo_sheet = _resolve_fcde_importo(source_wb)
        residui_titolo_i, residui_i_ref, residui_i_sheet = _resolve_fcde_residui_rendiconto(
            source_wb,
            "residuiattivifcde_titolo_uno",
        )

        residui_titolo_iii, residui_iii_ref, residui_iii_sheet = _resolve_fcde_residui_rendiconto(
            source_wb,
            "residuiattivifcde_titolo_tre",
        )
    finally:
        source_wb.close()

    source_parts = []
    if importo_ref:
        source_parts.append(f"Importo FCDE: {importo_sheet or 'sorgente'}, cella {importo_ref}")
    if residui_i_ref:
        source_parts.append(f"Residui titolo I: {residui_i_sheet or 'sorgente'}, cella {residui_i_ref}")
    if residui_iii_ref:
        source_parts.append(f"Residui titolo III: {residui_iii_sheet or 'sorgente'}, cella {residui_iii_ref}")

    if importo_fcde is not None:
        cell = ws_fcde.cell(row=FCDE_ROW_IMPORTO, column=year_col)
        cell.value = importo_fcde
        _set_comment(cell, source_file_name, f"foglio {importo_sheet or 'n/d'} | cella {importo_ref}", keep_source_reference)

    if residui_titolo_i is not None:
        cell = ws_fcde.cell(row=FCDE_ROW_RESIDUI_TITOLO_I, column=year_col)
        cell.value = residui_titolo_i
        _set_comment(cell, source_file_name, f"foglio {residui_i_sheet or 'n/d'} | cella {residui_i_ref}", keep_source_reference)

    if residui_titolo_iii is not None:
        cell = ws_fcde.cell(row=FCDE_ROW_RESIDUI_TITOLO_III, column=year_col)
        cell.value = residui_titolo_iii
        _set_comment(cell, source_file_name, f"foglio {residui_iii_sheet or 'n/d'} | cella {residui_iii_ref}", keep_source_reference)

    total_residui = None
    if residui_titolo_i is not None or residui_titolo_iii is not None:
        total_residui = Decimal("0")
        if residui_titolo_i is not None:
            total_residui += residui_titolo_i
        if residui_titolo_iii is not None:
            total_residui += residui_titolo_iii
        ws_fcde.cell(row=FCDE_ROW_TOTALE, column=year_col).value = total_residui

    percentage = None
    if importo_fcde is not None and total_residui not in (None, Decimal("0")):
        percentage = (importo_fcde * Decimal("100")) / total_residui

    if percentage is not None:
        cell = ws_fcde.cell(row=FCDE_ROW_PERCENTUALE, column=year_col)
        cell.value = percentage
        source_descr = "calcolata come Importo FCDE / Totale residui attivi EP * 100"
        _set_comment(cell, source_file_name, source_descr, keep_source_reference)

    ws_fcde.cell(row=FCDE_ROW_METODO, column=year_col).value = "ordinario"

    if log is not None:
        year_ref = get_column_letter(year_col)
        log(f"FCDE year {year}: populated column {year_ref}")

    return percentage, f"FCDE!{get_column_letter(year_col)}{FCDE_ROW_PERCENTUALE}"


def populate_fcde_sheet(
    analysis_path: Path,
    bdap_path: Optional[Path],
    year: int,
    keep_source_reference: bool = True,
    log: Optional[Callable[[str], None]] = None,
) -> tuple[Optional[Decimal], Optional[str]]:
    """Compila il foglio FCDE direttamente su disco."""
    wb_analysis = load_workbook(analysis_path)
    try:
        result = populate_fcde_sheet_in_workbook(
            wb_analysis,
            bdap_path,
            year,
            keep_source_reference=keep_source_reference,
            log=log,
        )
        wb_analysis.save(analysis_path)
        return result
    finally:
        wb_analysis.close()


def read_fcde_percentage_from_workbook(wb_analysis, year: int) -> tuple[Optional[Decimal], Optional[str]]:
    """Legge la percentuale FCDE già compilata nel workbook di output."""
    if FCDE_SHEET_NAME not in wb_analysis.sheetnames:
        return None, None
    return _read_fcde_percentage_from_sheet(wb_analysis[FCDE_SHEET_NAME], year)
