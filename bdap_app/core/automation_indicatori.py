"""Helper per scoprire e leggere il workbook degli indicatori."""

import re
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.comments import Comment

from ..support.text_utils import _coerce_year_value, normalize_text
from ..support.value_formatter import format_value_italian
from ..support.value_resolver import (
    find_bdap_value_by_label,
    find_bdap_value_with_fallback,
    format_source_location,
)


def _extract_years(path: Path) -> set[int]:
    text = f"{path.name} {path.stem}"
    return {int(match) for match in re.findall(r"(19\d{2}|20\d{2})", text)}


def _candidate_matches_token(candidate: Path, token: str) -> bool:
    candidate_text = normalize_text(str(candidate))
    if token in candidate_text:
        return True
    if token == "indicatori":
        return any(snippet in candidate_text for snippet in {"indicatori", "ind", "sintet", "sintetic"})
    return False


def _candidate_year_matches(path: Path, year: int) -> bool:
    if year in _extract_years(path):
        return True

    try:
        wb_temp = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False

    try:
        ws_temp = wb_temp.active
        for row_idx in range(1, 6):
            for col_idx in range(1, 4):
                if _coerce_year_value(ws_temp.cell(row=row_idx, column=col_idx).value) == year:
                    return True
    finally:
        wb_temp.close()

    return False


def _discover_aux_workbook(bdap_path: Optional[Path], token: str, year: int) -> Optional[Path]:
    """Localizza un workbook ausiliario coerente con l'anno richiesto."""
    if bdap_path is None:
        return None

    base_dir = bdap_path.parent
    if not base_dir.exists() or not base_dir.is_dir():
        return None

    search_roots = [base_dir]
    if normalize_text(base_dir.name) == "datibdap" and base_dir.parent.exists() and base_dir.parent.is_dir():
        search_roots.append(base_dir.parent)

    candidates = sorted(
        {
            candidate
            for root in search_roots
            for candidate in root.rglob("*.xlsx")
            if (candidate != bdap_path or token == "indicatori")
            and not candidate.name.startswith("~$")
            and _candidate_matches_token(candidate, token)
        }
    )
    if not candidates:
        return None

    for candidate in candidates:
        if _candidate_year_matches(candidate, year):
            return candidate

    if token == "indicatori":
        return None
    return candidates[0]


def _load_indicatori_workbook(bdap_path: Optional[Path], year: int) -> tuple[Optional[object], Optional[Path]]:
    """Carica il workbook indicatori se esiste per l'anno richiesto."""
    if bdap_path is None:
        return None, None

    try:
        path = _discover_aux_workbook(bdap_path, "indicatori", int(year))
        if path is None:
            return None, None
        return load_workbook(path, data_only=True), path
    except Exception:
        return None, None


def _resolve_indicatori_sheet(source_wb, data_source_mapping: dict):
    sheet_token = data_source_mapping.get("sheet_name_contains")
    if isinstance(sheet_token, str) and sheet_token.strip():
        token_norm = normalize_text(sheet_token)
        for name in source_wb.sheetnames:
            if token_norm in normalize_text(name):
                return source_wb[name]

    if "sheet_idx" not in data_source_mapping:
        return None

    try:
        idx = int(data_source_mapping["sheet_idx"])
    except Exception:
        return None

    if 1 <= idx <= len(source_wb.sheetnames):
        return source_wb[source_wb.sheetnames[idx - 1]]
    return None


def process_indicatori_row(
    target_cell,
    bdap_path: Optional[Path],
    year: int,
    special_key: Optional[str],
    data_source_mapping: dict,
    keep_source_reference: bool = True,
):
    """Risolve una riga con `source_workbook=indicatori`."""
    wb_indicatori, path = _load_indicatori_workbook(bdap_path, year)
    if wb_indicatori is None:
        if data_source_mapping.get("optional", False):
            target_cell.value = None
            target_cell.comment = None
            return False, 0, None, None, None, None
        raise KeyError("Missing indicatori workbook")

    ws_src = _resolve_indicatori_sheet(wb_indicatori, data_source_mapping)
    if ws_src is None:
        if data_source_mapping.get("optional", False):
            target_cell.value = None
            target_cell.comment = None
            return False, 0, wb_indicatori, None, path, None
        raise KeyError("Indicatori sheet not found")

    value = None
    actual_ref = None
    if "cell_ref" in data_source_mapping:
        value, actual_ref = find_bdap_value_with_fallback(ws_src, data_source_mapping["cell_ref"], max_offset=1)
    elif "label" in data_source_mapping:
        value, actual_ref = find_bdap_value_by_label(ws_src, data_source_mapping["label"])
    elif "cell_refs" in data_source_mapping:
        for cell_ref in data_source_mapping.get("cell_refs", []):
            value, actual_ref = find_bdap_value_with_fallback(ws_src, cell_ref, max_offset=1)
            if value is not None:
                break

    if value is None:
        if data_source_mapping.get("optional", False):
            target_cell.value = None
            target_cell.comment = None
            return False, 0, wb_indicatori, None, path, None
        raise KeyError("No value found in indicatori source")

    formatted = format_value_italian(value, percent=bool(data_source_mapping.get("is_percentage", False)))
    target_cell.value = formatted
    if keep_source_reference:
        source_file_name = path.name if path is not None else (bdap_path.name if bdap_path is not None else "n/d")
        target_cell.comment = Comment(
            f"Fonte indicatori: {source_file_name} | {format_source_location(ws_src.title, actual_ref)}",
            "BDAP Automation",
        )
    return True, 1, wb_indicatori, None, path, value
