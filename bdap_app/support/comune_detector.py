"""Modulo per rilevare il nome del comune in un workbook Excel."""
from pathlib import Path
from typing import Optional
from openpyxl import load_workbook
from .text_utils import normalize_text


def detect_comune_in_workbook(path: Path) -> Optional[str]:
    """Rileva il nome del comune in un workbook controllando A7 per 'Denominazione' e restituendo B7.

    Restituisce il valore stringa di B7 se A7 contiene un'etichetta come 'Denominazione Ente', altrimenti None.
    """
    wb = load_workbook(path, data_only=True)
    # controlla tutti i fogli, preferendo quello attivo poi gli altri
    sheets = [wb.active] + [s for s in wb.worksheets if s is not wb.active]
    for ws in sheets:
        try:
            a7 = ws.cell(row=7, column=1).value
            if a7 and "denominaz" in normalize_text(a7):
                b7 = ws.cell(row=7, column=2).value
                return str(b7).strip() if b7 is not None else None
        except Exception:
            continue
    return None


def detect_comune_from_path(path: Path) -> Optional[str]:
    """Dato un percorso file o cartella, rileva e restituisce il nome del comune quando possibile.

    Se `path` è una cartella, ricerca ricorsivamente il primo file .xlsx che contiene l'etichetta.
    """
    path = Path(path)
    if path.is_file():
        try:
            return detect_comune_in_workbook(path)
        except Exception:
            return None

    # Se è una cartella: cerca file .xlsx
    for p in sorted(path.rglob("*.xlsx")):
        try:
            result = detect_comune_in_workbook(p)
            if result:
                return result
        except Exception:
            continue
    return None
