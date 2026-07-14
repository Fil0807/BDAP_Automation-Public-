"""Utility per la risoluzione dei valori BDAP."""

import re
from typing import Optional
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
from .text_utils import label_match_score, normalize_text
from .value_formatter import coerce_numeric

NUMERIC_LIKE_RE = re.compile(r"[+-]?(?:\d{1,3}(?:[.\s]\d{3})+|\d+)(?:[.,]\d+)?\s*%?")


def is_empty_value(value) -> bool:
    """Verifica se il valore è None o una stringa vuota."""
    return value is None or (isinstance(value, str) and value.strip() == "")


def is_numeric_like(value: object) -> bool:
    """Verifica se il valore ha l'aspetto di un numero."""
    if isinstance(value, (int, float)):
        return True
    if not isinstance(value, str):
        return False
    return bool(NUMERIC_LIKE_RE.fullmatch(value.strip()))


def format_source_location(sheet_name: str, cell_ref: str) -> str:
    """Formatta un riferimento di cella in una descrizione leggibile."""
    if not cell_ref:
        return f"foglio {sheet_name}"
    try:
        row_idx, col_idx = coordinate_to_tuple(cell_ref)
    except Exception:
        return f"foglio {sheet_name}, cella {cell_ref}"
    return f"foglio {sheet_name}, riga {row_idx}, colonna {col_idx} ({cell_ref})"


def find_bdap_value_by_label(ws_bdap, label: str, label_cache: Optional[dict] = None):
    """Trova un valore nel foglio confrontando l'etichetta testuale."""
    target = normalize_text(label)
    if not target:
        return None, None

    cache_key = (id(ws_bdap), target)
    if label_cache is not None and cache_key in label_cache:
        return label_cache[cache_key]

    def _scan_row_for_value(row_idx: int, start_col: int = 1):
        for scan_col in range(start_col, ws_bdap.max_column + 1):
            candidate_cell = ws_bdap.cell(row=row_idx, column=scan_col)
            candidate = candidate_cell.value
            if is_empty_value(candidate):
                continue
            if is_numeric_like(candidate):
                return candidate, candidate_cell.coordinate
        return None, None

    def _collect_candidates(max_label_cols: int):
        found = []
        capped_cols = min(max_label_cols, ws_bdap.max_column)
        for row_idx in range(1, ws_bdap.max_row + 1):
            for col_idx in range(1, capped_cols + 1):
                cell = ws_bdap.cell(row=row_idx, column=col_idx)
                if not isinstance(cell.value, str):
                    continue
                score = label_match_score(label, cell.value)
                if score < 0.62:
                    continue
                found.append((score, row_idx, col_idx))
        return found

    candidates = _collect_candidates(max_label_cols=8)
    if not candidates and ws_bdap.max_column > 8:
        candidates = _collect_candidates(max_label_cols=ws_bdap.max_column)

    candidates.sort(key=lambda item: item[0], reverse=True)
    for _score, row_idx, col_idx in candidates:
        candidate_rows = [row_idx]
        for offset in range(1, 3):
            if row_idx - offset >= 1:
                candidate_rows.append(row_idx - offset)
            if row_idx + offset <= ws_bdap.max_row:
                candidate_rows.append(row_idx + offset)

        for candidate_row in candidate_rows:
            if candidate_row == row_idx:
                value, coordinate = _scan_row_for_value(candidate_row, start_col=col_idx + 1)
                if value is not None:
                    result = (value, coordinate)
                    if label_cache is not None:
                        label_cache[cache_key] = result
                    return result
                value, coordinate = _scan_row_for_value(candidate_row, start_col=1)
                if value is not None:
                    result = (value, coordinate)
                    if label_cache is not None:
                        label_cache[cache_key] = result
                    return result
                continue

            value, coordinate = _scan_row_for_value(candidate_row, start_col=1)
            if value is not None:
                result = (value, coordinate)
                if label_cache is not None:
                    label_cache[cache_key] = result
                return result

    result = (None, None)
    if label_cache is not None:
        label_cache[cache_key] = result
    return result


def find_bdap_value_in_workbook_by_label(wb_bdap, label: str, label_cache: Optional[dict] = None):
    """Cerca nel workbook l'etichetta e restituisce il primo valore corrispondente."""
    for ws in wb_bdap.worksheets:
        value, actual_cell_ref = find_bdap_value_by_label(ws, label, label_cache=label_cache)
        if value is not None:
            return value, format_source_location(ws.title, actual_cell_ref)
    return None, None


def find_bdap_value_with_fallback(ws_bdap, base_cell_ref: str, max_offset: int = 2):
    """Trova il valore a partire da una cella, con fallback sulle righe adiacenti."""
    try:
        row, col = coordinate_to_tuple(base_cell_ref)
    except Exception:
        return None, None
    
    try:
        cell = ws_bdap.cell(row=row, column=col)
        if not is_empty_value(cell.value):
            if isinstance(cell.value, str):
                if not NUMERIC_LIKE_RE.fullmatch(cell.value.strip()):
                    raise ValueError("Reference cell contains text, searching nearby rows")
            return cell.value, cell.coordinate
    except Exception:
        pass
    
    for offset in range(1, max_offset + 1):
        for delta in [-offset, offset]:
            try:
                new_row = row + delta
                if new_row < 1 or new_row > ws_bdap.max_row:
                    continue
                cell = ws_bdap.cell(row=new_row, column=col)
                if not is_empty_value(cell.value):
                    if isinstance(cell.value, str) and not NUMERIC_LIKE_RE.fullmatch(cell.value.strip()):
                        continue
                    adjusted_ref = f"{get_column_letter(col)}{new_row}"
                    return cell.value, adjusted_ref
            except Exception:
                continue
    
    return None, None


def _find_row_by_expected_label_near_cell(
    ws_bdap,
    base_cell_ref: str,
    expected_label: str,
    max_row_offset: int,
    max_label_columns: int,
    label_check_cells: Optional[list[str]],
    strict_expected_label: bool,
    search_key_prefix: str | None,
) -> Optional[int]:
    try:
        base_row, _ = coordinate_to_tuple(base_cell_ref)
    except Exception:
        return None

    expected_norm = normalize_text(expected_label)

    def match_label(label_text: str) -> float:
        label_norm = normalize_text(label_text)
        if label_norm == expected_norm:
            return 1.0
        if strict_expected_label:
            return 0.0
        return label_match_score(expected_label, label_text)

    if search_key_prefix:
        for row_idx in range(1, ws_bdap.max_row + 1):
            cell_value = ws_bdap.cell(row=row_idx, column=1).value
            if isinstance(cell_value, str) and cell_value.strip().startswith(search_key_prefix):
                return row_idx
        return None

    if label_check_cells:
        label_cells = label_check_cells if isinstance(label_check_cells, (list, tuple)) else [label_check_cells]
        for label_cell_ref in label_cells:
            try:
                if ":" in str(label_cell_ref):
                    label_cell_ref = str(label_cell_ref).split(":")[0]
                anchor_row, anchor_col = coordinate_to_tuple(label_cell_ref)
            except Exception:
                continue

            offsets = [0]
            for offset in range(1, max_row_offset + 1):
                offsets.extend([-offset, offset])

            for offset in offsets:
                row_idx = anchor_row + offset
                if row_idx < 1 or row_idx > ws_bdap.max_row:
                    continue

                first_col = ws_bdap.cell(row=row_idx, column=1).value
                if isinstance(first_col, str) and match_label(first_col) >= 0.84:
                    return row_idx

                anchor_value = ws_bdap.cell(row=row_idx, column=anchor_col).value
                if isinstance(anchor_value, str) and match_label(anchor_value) >= 0.84:
                    return row_idx

    best = None
    row_start = max(1, base_row - max_row_offset)
    row_end = min(ws_bdap.max_row, base_row + max_row_offset)

    for row_idx in range(row_start, row_end + 1):
        label_text = ""
        first_col = ws_bdap.cell(row=row_idx, column=1).value
        if isinstance(first_col, str) and first_col.strip():
            label_text = first_col.strip()
        else:
            for col_idx in range(2, min(max_label_columns, ws_bdap.max_column) + 1):
                cell_value = ws_bdap.cell(row=row_idx, column=col_idx).value
                if isinstance(cell_value, str) and cell_value.strip():
                    label_text = cell_value.strip()
                    break
        if not label_text:
            continue

        score = match_label(label_text)
        if score < 0.84:
            continue

        candidate = (score, row_idx)
        if best is None or candidate[0] > best[0]:
            best = candidate

    if best is None:
        return None
    return best[1]


def find_bdap_value_by_expected_label_near_cell(
    ws_bdap,
    base_cell_ref: str,
    expected_label: str,
    max_row_offset: int = 2,
    max_label_columns: int = 4,
    label_check_cells: Optional[list[str]] = None,
    strict_expected_label: bool = False,
    search_key_prefix: str = None,
):
    """Risolve il valore vicino a una cella fissa solo se il titolo di riga vicino corrisponde all'etichetta attesa."""
    if not expected_label:
        return None, None
    try:
        _, base_col = coordinate_to_tuple(base_cell_ref)
    except Exception:
        return None, None

    def _value_from_row(row_idx: int):
        value_cell = ws_bdap.cell(row=row_idx, column=base_col)
        value = value_cell.value
        if is_empty_value(value):
            return None, None
        if isinstance(value, str) and not is_numeric_like(value):
            return None, None
        return value, value_cell.coordinate

    row_idx = _find_row_by_expected_label_near_cell(
        ws_bdap,
        base_cell_ref,
        expected_label=expected_label,
        max_row_offset=max_row_offset,
        max_label_columns=max_label_columns,
        label_check_cells=label_check_cells,
        strict_expected_label=strict_expected_label,
        search_key_prefix=search_key_prefix,
    )
    if row_idx is None:
        return None, None
    return _value_from_row(row_idx)


def resolve_row_value_from_label(wb_bdap, row_label: str, sheet_hint: Optional[str] = None, label_cache: Optional[dict] = None):
    """Risolve il valore di riga per etichetta, preferendo il foglio suggerito."""
    if not row_label:
        return None, None

    if sheet_hint is not None:
        try:
            idx = int(float(str(sheet_hint).strip()))
            ws_hint = wb_bdap[wb_bdap.sheetnames[idx - 1]]
            value, source = find_bdap_value_by_label(ws_hint, row_label, label_cache=label_cache)
            if value is not None:
                return value, source
        except Exception:
            pass

    return find_bdap_value_in_workbook_by_label(wb_bdap, row_label, label_cache=label_cache)
