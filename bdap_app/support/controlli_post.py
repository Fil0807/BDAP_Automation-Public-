"""Lettura e riepilogo delle criticita dai file Controlli Post."""

from __future__ import annotations

from collections import Counter, defaultdict
from copy import copy
from dataclasses import dataclass
from pathlib import Path
import re
import unicodedata
from typing import Iterable

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill


REPORT_SHEET_NAME = "CONTROLLI-POST"
_SEVERITY_CRITICAL = "Criticità contabile"
_SEVERITY_WARNING = "Warning"
_REPORT_COLUMNS = 7
_CRITICAL_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
_WARNING_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")


@dataclass(frozen=True)
class ControlloPostIssue:
    year: int | None
    source_path: Path
    row: int
    control_id: str
    area: str
    severity: str
    description: str


@dataclass(frozen=True)
class ControlliPostReport:
    files: tuple[Path, ...]
    issues: tuple[ControlloPostIssue, ...]

    @property
    def has_files(self) -> bool:
        return bool(self.files)

    @property
    def has_issues(self) -> bool:
        return bool(self.issues)

    def counts_by_severity(self) -> Counter[str]:
        return Counter(issue.severity for issue in self.issues)

    def counts_by_year_and_severity(self) -> dict[int | None, Counter[str]]:
        counts: dict[int | None, Counter[str]] = defaultdict(Counter)
        for issue in self.issues:
            counts[issue.year][issue.severity] += 1
        return dict(counts)


def _fold_text(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text.lower()).strip()


def _display_text(value: object) -> str:
    return re.sub(r"[ \t\r\f\v]+", " ", str(value or "")).strip()


def _one_line(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _years_from_name(path: Path) -> set[int]:
    return {int(match) for match in re.findall(r"\b(19\d{2}|20\d{2}|2100)\b", path.name)}


def discover_controlli_post_files(base_dir: Path, years: Iterable[int] | None = None) -> list[Path]:
    """Trova i file Excel Controlli Post sotto la cartella del comune."""
    year_filter = set(int(year) for year in years or ())
    files: list[Path] = []
    if not base_dir.exists():
        return files

    for path in base_dir.rglob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        folded_name = _fold_text(path.name)
        if "controlli" not in folded_name or "post" not in folded_name:
            continue
        file_years = _years_from_name(path)
        if year_filter and file_years and file_years.isdisjoint(year_filter):
            continue
        files.append(path)

    return sorted(files, key=lambda item: (min(_years_from_name(item) or {9999}), item.name.lower()))


def _severity_from_value(value: object) -> str:
    folded = _fold_text(value)
    if not folded:
        return ""
    if folded == "warning":
        return _SEVERITY_WARNING
    if folded in {"criticita", "criticita contabile", "criticita contabili"}:
        return _SEVERITY_CRITICAL
    return ""


def _split_nonempty_lines(value: object) -> list[str]:
    return [_display_text(line) for line in str(value or "").splitlines() if _display_text(line)]


def _severity_from_control_text(value: object) -> str:
    lines = _split_nonempty_lines(value)
    if not lines:
        return ""
    return _severity_from_value(lines[-1])


def _description_without_trailing_severity(value: object) -> str:
    lines = _split_nonempty_lines(value)
    if lines and _severity_from_value(lines[-1]):
        lines = lines[:-1]
    return "\n".join(lines)


def _find_controlli_sheet(workbook) -> object | None:
    for sheet in workbook.worksheets:
        folded = _fold_text(sheet.title)
        if "controlli" in folded and "post" in folded:
            return sheet
    return None


def _find_output_sheet(workbook) -> object | None:
    target = _fold_text(REPORT_SHEET_NAME)
    for sheet in workbook.worksheets:
        if _fold_text(sheet.title) == target:
            return sheet
    return None


def _find_header_columns(ws) -> tuple[int, dict[str, int]]:
    for row in range(1, min(ws.max_row or 0, 10) + 1):
        headers = {
            _fold_text(ws.cell(row=row, column=col).value): col
            for col in range(1, min(ws.max_column or 0, 12) + 1)
        }
        control_col = next((col for header, col in headers.items() if "controllo" in header), None)
        if control_col is None:
            continue
        columns = {
            "id": next((col for header, col in headers.items() if header == "id"), 1),
            "area": next((col for header, col in headers.items() if "area" in header), 2),
            "control": control_col,
            "outcome": next((col for header, col in headers.items() if "esito" in header), 4),
        }
        return row, columns
    return 1, {"id": 1, "area": 2, "control": 3, "outcome": 4}


def read_controlli_post_issues(path: Path) -> list[ControlloPostIssue]:
    """Estrae solo le righe con esito Warning/Criticita dal foglio Controlli Post."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = _find_controlli_sheet(workbook)
        if ws is None:
            return []

        header_row, columns = _find_header_columns(ws)
        file_years = sorted(_years_from_name(path))
        file_year = file_years[-1] if file_years else None
        issues: list[ControlloPostIssue] = []

        for row in range(header_row + 1, (ws.max_row or 0) + 1):
            area = _display_text(ws.cell(row=row, column=columns["area"]).value)
            control = ws.cell(row=row, column=columns["control"]).value
            outcome = ws.cell(row=row, column=columns["outcome"]).value
            severity = _severity_from_value(outcome) or _severity_from_control_text(control)
            if not severity:
                continue

            description = _description_without_trailing_severity(control)
            if "conteggio" in _fold_text(area) or "conteggio" in _fold_text(description):
                continue

            control_id = _display_text(ws.cell(row=row, column=columns["id"]).value)
            issues.append(
                ControlloPostIssue(
                    year=file_year,
                    source_path=path,
                    row=row,
                    control_id=control_id,
                    area=area,
                    severity=severity,
                    description=description,
                )
            )
        return issues
    finally:
        workbook.close()


def collect_controlli_post_report(
    base_dir: Path,
    years: Iterable[int] | None = None,
) -> ControlliPostReport:
    files = discover_controlli_post_files(base_dir, years=years)
    issues: list[ControlloPostIssue] = []
    for path in files:
        issues.extend(read_controlli_post_issues(path))
    return ControlliPostReport(
        files=tuple(files),
        issues=tuple(issues),
    )


def _copy_row_style(ws, source_row: int, target_row: int) -> None:
    if source_row == target_row:
        return
    for col_idx in range(1, _REPORT_COLUMNS + 1):
        source_cell = ws.cell(row=source_row, column=col_idx)
        target_cell = ws.cell(row=target_row, column=col_idx)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def _find_output_header_row(ws) -> int:
    for row_idx in range(1, min(ws.max_row or 0, 20) + 1):
        row_headers = [_fold_text(ws.cell(row=row_idx, column=col_idx).value) for col_idx in range(1, _REPORT_COLUMNS + 1)]
        if "anno" in row_headers and "livello" in row_headers and any("controllo" in header for header in row_headers):
            return row_idx
    return 5


def _clear_report_data(ws, data_start_row: int) -> None:
    max_row = max(ws.max_row or data_start_row, data_start_row)
    for row_idx in range(data_start_row, max_row + 1):
        for col_idx in range(1, _REPORT_COLUMNS + 1):
            ws.cell(row=row_idx, column=col_idx).value = None


def _fill_for_severity(severity: str) -> PatternFill | None:
    if severity == _SEVERITY_CRITICAL:
        return _CRITICAL_FILL
    if severity == _SEVERITY_WARNING:
        return _WARNING_FILL
    return None


def _apply_issue_row_fill(ws, row_idx: int, severity: str) -> None:
    fill = _fill_for_severity(severity)
    if fill is None:
        return
    for col_idx in range(1, _REPORT_COLUMNS + 1):
        ws.cell(row=row_idx, column=col_idx).fill = copy(fill)


def write_controlli_post_sheet(output_path: Path, report: ControlliPostReport) -> bool:
    """Compila il foglio CONTROLLI-POST gia presente nel template."""
    if not report.has_files:
        return False

    workbook: Workbook = load_workbook(output_path)
    try:
        ws = _find_output_sheet(workbook)
        if ws is None:
            return False

        header_row = _find_output_header_row(ws)
        data_start_row = header_row + 1
        _clear_report_data(ws, data_start_row)
        counts = report.counts_by_severity()
        ws["A2"] = f"File analizzati: {len(report.files)}"
        ws["A3"] = (
            f"Segnalazioni: {len(report.issues)}; "
            f"Criticità contabili: {counts.get(_SEVERITY_CRITICAL, 0)}; "
            f"Warning: {counts.get(_SEVERITY_WARNING, 0)}"
        )

        if report.has_issues:
            for row_idx, issue in enumerate(report.issues, start=data_start_row):
                _copy_row_style(ws, data_start_row, row_idx)
                values = [
                    issue.year,
                    issue.severity,
                    issue.area,
                    issue.control_id,
                    issue.description,
                    issue.source_path.name,
                    issue.row,
                ]
                for col_idx, value in enumerate(values, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                _apply_issue_row_fill(ws, row_idx, issue.severity)
        else:
            ws.cell(
                row=data_start_row,
                column=1,
                value="Nessuna criticità o warning segnalati nei file Controlli Post analizzati.",
            )

        workbook.save(output_path)
        return True
    finally:
        workbook.close()


def format_controlli_post_summary(report: ControlliPostReport, max_items: int = 8) -> str:
    if not report.has_files:
        return "Nessun file Controlli Post trovato."

    lines = ["Controlli Post analizzati:"]
    for year, counts in sorted(report.counts_by_year_and_severity().items(), key=lambda item: item[0] or 0):
        label = str(year) if year is not None else "anno non rilevato"
        lines.append(
            f"- {label}: "
            f"{counts.get(_SEVERITY_CRITICAL, 0)} criticità contabili, "
            f"{counts.get(_SEVERITY_WARNING, 0)} warning"
        )

    if not report.has_issues:
        lines.append("Nessuna criticità o warning segnalati.")
        return "\n".join(lines)

    shown_items = max(0, max_items)
    if shown_items:
        lines.append("")
        lines.append("Prime segnalazioni:")
        for issue in report.issues[:shown_items]:
            year = issue.year if issue.year is not None else "n/d"
            description = _one_line(issue.description)
            if len(description) > 150:
                description = description[:147].rstrip() + "..."
            lines.append(f"- {year} | {issue.severity} | {issue.area}: {description}")
    remaining = len(report.issues) - shown_items
    if remaining > 0:
        lines.append(f"... altre {remaining} segnalazioni nel foglio '{REPORT_SHEET_NAME}'.")
    return "\n".join(lines)
