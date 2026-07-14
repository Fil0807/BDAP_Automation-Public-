"""Helper per la scoperta di percorsi e dataset per i workflow CLI di BDAP."""

from pathlib import Path
import re
from typing import Optional
from openpyxl import load_workbook
import sys

# Individua file con nome tipo "Rend. YYYY.xlsx" (case-insensitive, spazi consentiti).
BDAP_YEAR_FILE_PATTERN = re.compile(r"Rend\.\s*(\d{4})\.xlsx$", re.IGNORECASE)
GENERIC_YEAR_PATTERN = re.compile(r"(19\d{2}|20\d{2})")
GENERIC_TWO_DIGIT_YEAR_PATTERN = re.compile(r"(?<!\d)(\d{2})(?!\d)")

__all__ = [
    "BDAP_YEAR_FILE_PATTERN",
    "GENERIC_YEAR_PATTERN",
    "normalize_token",
    "dedupe_paths",
    "infer_new_analysis_path",
    "infer_bdap_dir",
    "discover_comune_base_dirs",
    "resolve_base_dir_for_comune",
    "discover_bdap_files_by_year",
    "resolve_default_template_path",
    "_is_flat_bdap_folder",
    "infer_comune_name_from_bdap_dir",
]

COMUNE_INLINE_PATTERN = re.compile(r"\bcomune\s+di\b\s*[:\-]?\s*(.+)", re.IGNORECASE)
COMUNE_LABEL_PATTERN = re.compile(r"^\s*comune\s+di\s*[:\-]?\s*$", re.IGNORECASE)
GENERIC_COMUNE_VALUE_TOKENS = {
    "nome",
    "ente",
    "denominazione",
    "denominazioneente",
    "comune",
}


def _parse_two_digit_year_from_stem(stem: str) -> Optional[int]:
    # Accetta token anno corto (YY) solo per nomi file BDAP-like e lo mappa a 20YY.
    # Esempio: "Rend. 26.xlsx" -> 2026.
    lower_stem = stem.lower()
    if not any(keyword in lower_stem for keyword in ("rend", "rendiconto", "bdap")):
        return None

    matches = GENERIC_TWO_DIGIT_YEAR_PATTERN.findall(stem)
    if len(matches) != 1:
        return None

    yy = int(matches[0])
    return 2000 + yy


def normalize_token(value: str) -> str:
    # Normalizza etichette per un matching robusto dei nomi di cartelle.
    return re.sub(r"[^a-z0-9]", "", value.lower())


EXCLUDED_BDAP_YEAR_FILE_TOKENS = {
    "analisi",
    "compilato",
    "controllipost",
    "debiti",
    "questionario",
    "template",
}


def _bdap_year_file_priority(path: Path) -> Optional[int]:
    """Classifica i soli file annuali BDAP, escludendo workbook ausiliari."""
    normalized_stem = normalize_token(path.stem)
    if any(token in normalized_stem for token in EXCLUDED_BDAP_YEAR_FILE_TOKENS):
        return None
    if BDAP_YEAR_FILE_PATTERN.search(path.name):
        return 0
    if "rendiconto" in normalized_stem:
        return 1
    if "rend" in normalized_stem:
        return 2
    if "bdap" in normalized_stem:
        return 3
    return None


def dedupe_paths(paths: list[Path]) -> list[Path]:
    # Mantiene l'ordine della prima occorrenza rimuovendo duplicati di percorso.
    unique: list[Path] = []
    seen: set[str] = set()
    for path in paths:
        try:
            key = str(path.expanduser().resolve())
        except OSError:
            key = str(path.expanduser())
        if key not in seen:
            unique.append(path)
            seen.add(key)
    return unique


def _has_bdap_year_files(folder: Path) -> bool:
    # Restituisce True se la cartella contiene direttamente file BDAP per anno.
    # Accetta sia nomi stretti (Rend. YYYY.xlsx) che varianti con anno a 4 cifre.
    if not folder.exists() or not folder.is_dir():
        return False
    return bool(discover_bdap_files_by_year(folder))


def _is_flat_bdap_folder(folder: Path) -> bool:
    # Restituisce True se la cartella contiene Rend. YYYY.xlsx direttamente senza struttura "1 doc".
    if not _has_bdap_year_files(folder):
        return False
    # Check if it has the proper comune structure (1 doc subfolder).
    if (folder / "1 doc").exists() or any(
        p.is_dir() and normalize_token(p.name) in {"1doc", "doc1", "1documento", "documento1"}
        for p in folder.iterdir()
    ):
        return False
    return True


def _clean_comune_name(raw: str) -> str:
    # Pulisce le etichette estratte rimuovendo punteggiatura e rumore comune.
    cleaned = re.sub(r"\s+", " ", raw).strip(" .:-\t\n\r")
    return cleaned


def _looks_like_real_comune_name(value: str) -> bool:
    token = normalize_token(value)
    if not token:
        return False
    if token in GENERIC_COMUNE_VALUE_TOKENS:
        return False
    if len(token) < 3:
        return False
    return True


def _extract_comune_from_workbook(path: Path) -> Optional[str]:
    # Estrae l'etichetta del comune scansionando l'area in alto a sinistra del workbook per 'COMUNE DI'.
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return None

    try:
        for sheet in wb.worksheets[:3]:
            # Nei workbook Rend YYYY il nome del comune spesso è in B7, accanto ad A7 = "Denominazione ente".
            try:
                a7_value = sheet["A7"].value
                b7_value = sheet["B7"].value
            except Exception:
                a7_value = None
                b7_value = None

            if isinstance(a7_value, str) and normalize_token(a7_value) == normalize_token("Denominazione ente"):
                # Scansiona la riga a destra di A7 e conserva il primo valore che sembra un nome di comune reale.
                for col_idx in range(2, min(sheet.max_column or 0, 12) + 1):
                    row_value = sheet.cell(row=7, column=col_idx).value
                    if not isinstance(row_value, str):
                        continue

                    candidate = _clean_comune_name(row_value)
                    if not candidate:
                        continue

                    inline_match = COMUNE_INLINE_PATTERN.search(candidate)
                    if inline_match:
                        candidate = _clean_comune_name(inline_match.group(1))

                    if _looks_like_real_comune_name(candidate):
                        return candidate

            max_row = min(sheet.max_row or 0, 40)
            max_col = min(sheet.max_column or 0, 12)
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    value = sheet.cell(row=row_idx, column=col_idx).value
                    if not isinstance(value, str):
                        continue

                    text = value.strip()
                    if not text:
                        continue

                    inline_match = COMUNE_INLINE_PATTERN.search(text)
                    if inline_match:
                        candidate = _clean_comune_name(inline_match.group(1))
                        if candidate:
                            return candidate

                    if COMUNE_LABEL_PATTERN.match(text):
                        right_value = None
                        if col_idx + 1 <= max_col:
                            right_value = sheet.cell(row=row_idx, column=col_idx + 1).value
                        if isinstance(right_value, str):
                            candidate = _clean_comune_name(right_value)
                            if candidate:
                                return candidate
    finally:
        wb.close()

    return None


def infer_comune_name_from_bdap_dir(folder: Path) -> Optional[str]:
    # Inferisce il nome del comune dai workbook BDAP annuali quando è selezionata solo una cartella dati BDAP piatta.
    if not _is_flat_bdap_folder(folder):
        return None

    # Use discover_bdap_files_by_year to get only actual BDAP files (not analysis files)
    bdap_files = discover_bdap_files_by_year(folder)
    for year in sorted(bdap_files.keys())[:3]:  # Try first 3 years
        workbook_path = bdap_files[year]
        comune_name = _extract_comune_from_workbook(workbook_path)
        if comune_name:
            return comune_name

    # Fallback: prova a usare il nome della cartella padre se sembra un comune
    parent_name = folder.parent.name
    parent_normalized = normalize_token(parent_name)
    if parent_name and parent_normalized not in {"downloads", "desktop", "documents", "dati", "datibdap"}:
        return parent_name

    return None


def infer_new_analysis_path(base_dir: Path) -> Path:
    # Risolve il file di analisi "Nuovo Modello" più probabile nella directory base.
    preferred = base_dir / "Analisi Nuovo Modello.xlsx"
    if preferred.exists():
        return preferred

    preferred_subdir = base_dir / "2 analisi funzionaria"
    if preferred_subdir.exists():
        subdir_candidates = sorted(
            p for p in preferred_subdir.glob("Analisi*.xlsx")
            if not p.name.upper().startswith("REV_")
        )
        if subdir_candidates:
            return subdir_candidates[0]

    candidates = sorted(base_dir.glob("Analisi*Nuovo*Modello*.xlsx"))
    if candidates:
        return candidates[0]

    fallback_candidates = sorted(
        p  for p in base_dir.glob("Analisi*.xlsx")
        if not p.name.upper().startswith("REV_")
    )
    if fallback_candidates:
        return fallback_candidates[0]

    # Fallback: search in parent directory if base_dir is a flat BDAP folder
    parent_dir = base_dir.parent
    if parent_dir.exists() and parent_dir != base_dir:
        preferred_parent = parent_dir / "Analisi Nuovo Modello.xlsx"
        if preferred_parent.exists():
            return preferred_parent
        
        parent_candidates = sorted(
            p for p in parent_dir.glob("Analisi*.xlsx")
            if not p.name.upper().startswith("REV_")
        )
        if parent_candidates:
            return parent_candidates[0]

    return preferred


def infer_bdap_dir(base_dir: Path) -> Path:
    # Individua la directory di input BDAP secondo la struttura prevista o varianti normalizzate.
    preferred = base_dir / "1 doc" / "dati BDAP"
    if preferred.exists():
        return preferred

    if not base_dir.exists() or not base_dir.is_dir():
        return preferred

    # Support flat datasets where Rend. YYYY.xlsx files are stored directly in base_dir.
    if _has_bdap_year_files(base_dir):
        return base_dir

    doc_dir = base_dir / "1 doc"
    doc_dir_candidates: list[Path] = []
    if doc_dir.exists() and doc_dir.is_dir():
        doc_dir_candidates.append(doc_dir)
    else:
        for candidate in sorted(p for p in base_dir.iterdir() if p.is_dir()):
            token = normalize_token(candidate.name)
            if token in {"1doc", "doc1", "1documento", "documento1"}:
                doc_dir_candidates.append(candidate)

    for doc_candidate in doc_dir_candidates:
        for candidate in sorted(p for p in doc_candidate.iterdir() if p.is_dir()):
            if normalize_token(candidate.name) == "datibdap":
                return candidate

    return preferred


def discover_comune_base_dirs(workspace_root: Path) -> list[Path]:
    # Scopre le radici dei dataset che espongono una directory BDAP valida.
    candidates: list[Path] = []
    if not workspace_root.exists() or not workspace_root.is_dir():
        return candidates

    # Accept roots that already point to a dataset folder (comune level).
    workspace_bdap_dir = infer_bdap_dir(workspace_root)
    if workspace_bdap_dir.exists() and workspace_bdap_dir.is_dir():
        candidates.append(workspace_root)

    normalized_root_name = normalize_token(workspace_root.name)

    # Accept selecting ".../<comune>/1 doc" directly.
    if normalized_root_name in {"1doc", "doc1", "1documento", "documento1"}:
        comune_dir = workspace_root.parent
        bdap_dir = infer_bdap_dir(comune_dir)
        if bdap_dir.exists() and bdap_dir.is_dir():
            candidates.append(comune_dir)

    # Accept selecting ".../<comune>/1 doc/dati BDAP" directly.
    if normalized_root_name == "datibdap":
        # First try: is this part of a proper <comune>/1 doc/dati BDAP structure?
        doc_dir = workspace_root.parent
        comune_dir = doc_dir.parent
        if doc_dir.exists() and doc_dir.is_dir() and comune_dir.exists() and comune_dir.is_dir():
            if normalize_token(doc_dir.name) in {"1doc", "doc1", "1documento", "documento1"}:
                candidates.append(comune_dir)
        # Second try: is this a flat "dati BDAP" folder with BDAP files directly inside?
        elif _has_bdap_year_files(workspace_root):
            # Treat the "dati BDAP" folder itself as a comune
            candidates.append(workspace_root)

    for comune_dir in sorted(p for p in workspace_root.iterdir() if p.is_dir()):
        bdap_dir = infer_bdap_dir(comune_dir)
        if bdap_dir.exists() and bdap_dir.is_dir():
            candidates.append(comune_dir)

    return dedupe_paths(candidates)


def resolve_base_dir_for_comune(workspace_root: Path, comune: str) -> Path:
    # Risolve una radice dataset dal nome del comune con controlli di ambiguità.
    normalized_target = normalize_token(comune)
    if not normalized_target:
        raise ValueError("--comune cannot be empty")

    candidates = discover_comune_base_dirs(workspace_root)
    if not candidates:
        raise FileNotFoundError(
            f"No dataset folder found under {workspace_root}. "
            "Expected pattern: <folder>/1 doc/dati BDAP. "
            "You can pass either: workspace root with many comuni, a single comune folder, "
            "the '1 doc' folder, the 'dati BDAP' folder, or a folder containing BDAP .xlsx files with year in filename (YYYY or YY->20YY for BDAP-like names)."
        )

    matched: list[Path] = []
    exact_name_matches: list[Path] = []  # Prefer exact folder name matches
    inferred_name_matches: list[Path] = []  # Fallback to inferred name matches
    
    for path in candidates:
        folder_name_normalized = normalize_token(path.name)
        inferred_name = infer_comune_name_from_bdap_dir(path)
        inferred_name_normalized = normalize_token(inferred_name) if inferred_name else None
        
        # Check if target matches folder name (exact match)
        if normalized_target in folder_name_normalized or folder_name_normalized == normalized_target:
            exact_name_matches.append(path)
            matched.append(path)
        # Check if target matches inferred name (fallback)
        elif inferred_name_normalized and (normalized_target in inferred_name_normalized or inferred_name_normalized == normalized_target):
            inferred_name_matches.append(path)
            matched.append(path)
    
    # Prioritize exact folder name matches
    if exact_name_matches:
        if len(exact_name_matches) == 1:
            return exact_name_matches[0]
        # Multiple exact matches: this is truly ambiguous
        options = "\n".join(f"- {p.name}" for p in exact_name_matches)
        raise ValueError(
            f"Comune '{comune}' is ambiguous. Matching folders:\n{options}"
        )
    
    # Fallback to inferred name matches
    if inferred_name_matches:
        if len(inferred_name_matches) == 1:
            return inferred_name_matches[0]
        # Multiple inferred matches: also ambiguous
        options = "\n".join(f"- {p.name}" for p in inferred_name_matches)
        raise ValueError(
            f"Comune '{comune}' is ambiguous. Matching folders:\n{options}"
        )
    
    if len(matched) == 1:
        return matched[0]
    if len(matched) > 1:
        options = "\n".join(f"- {p.name}" for p in matched)
        raise ValueError(
            f"Comune '{comune}' is ambiguous. Matching folders:\n{options}"
        )

    # No matches: show all options available
    all_options = "\n".join(f"- {p.name}" for p in candidates)
    raise FileNotFoundError(
        f"No dataset folder matched comune '{comune}'. Available folders:\n{all_options}"
    )


def discover_bdap_files_by_year(bdap_dir: Path) -> dict[int, Path]:
    # Costruisce una mappa: anno -> percorso file BDAP
    best_by_year: dict[int, tuple[int, str, Path]] = {}
    if not bdap_dir.exists():
        return {}

    for path in sorted(bdap_dir.glob("*.xlsx")):
        priority = _bdap_year_file_priority(path)
        if priority is None:
            continue

        match = BDAP_YEAR_FILE_PATTERN.search(path.name)
        year: Optional[int] = None
        if match:
            # Extract YYYY year from filename pattern like "Rend. YYYY.xlsx"
            year = int(match.group(1))
        else:
            generic = GENERIC_YEAR_PATTERN.search(path.stem)
            if generic:
                year = int(generic.group(1))
            else:
                year = _parse_two_digit_year_from_stem(path.stem)
        if year is not None:
            candidate = (priority, path.name.lower(), path)
            if year not in best_by_year or candidate < best_by_year[year]:
                best_by_year[year] = candidate
    return {year: item[2] for year, item in sorted(best_by_year.items())}


def resolve_default_template_path(
    workspace_root: Path,
    project_root: Optional[Path] = None,
) -> Optional[Path]:
    """Resolve default template path using both project and workspace roots."""
    roots: list[Path] = [workspace_root]
    
    resource_root = get_resource_root()
    roots.insert(0, resource_root)

    if project_root is None:
        roots.insert(0, project_root)
    
    # Remove duplicate paths while preserving order.
    search_roots = dedupe_paths(roots)

    preferred_names = (
        "Template_Analisi.xlsx",
        "Template Analisi Contabile.xlsx",
        "Template_Analisi_Contabile.xlsx",
    )

    for root in search_roots:
        for name in preferred_names:
            candidate = root / name
            if candidate.exists() and candidate.is_file():
                return candidate

    for root in search_roots:
        candidates = sorted(root.glob("Template*Analisi*Contabile*.xlsx"))
        if candidates:
            return candidates[0]

    # If only a Google Sheets shortcut exists, suggest exported .xlsx path.
    for root in search_roots:
        gsheet_candidates = sorted(root.glob("Template*Analisi*Contabile*.gsheet"))
        if gsheet_candidates:
            return gsheet_candidates[0].with_suffix(".xlsx")

    return None


def get_resource_root()-> Path:
    """
    Restituisce la cartella contenente le risorse dell'applicazione.
    """

    # PyInstaller
    if hasattr(sys, "_MEIPASS"):
        return Path(sys._MEIPASS)
    
    # p2app
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent.parent / "Resources"
    
    return Path(__file__).resolve().parents[2]