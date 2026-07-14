"""Helper per l'esecuzione della pipeline nei workflow CLI di BDAP."""

from pathlib import Path
import argparse
import shutil
from typing import Callable, Optional

from openpyxl import load_workbook

from ..core.automation_fcde import populate_fcde_sheet
from ..core.automation import fill_accounting_analysis
from ..core.automation_search import find_sheet, find_year_column
from ..support.controlli_post import (
    collect_controlli_post_report,
    format_controlli_post_summary,
    write_controlli_post_sheet,
)
from ..support.text_utils import _coerce_year_value
from .cli_discovery import (
    discover_bdap_files_by_year,
    infer_bdap_dir,
    infer_new_analysis_path,
    resolve_base_dir_for_comune,
)

TEMPLATE_YEAR_START = 2020
TEMPLATE_YEAR_END = 2026

__all__ = [
    "TEMPLATE_YEAR_START",
    "TEMPLATE_YEAR_END",
    "parse_years_arg",
    "validate_inputs",
    "apply_comune_defaults",
    "ensure_analysis_input",
    "resolve_years_to_process",
    "run_pipeline",
]


# Analizza una lista separata da virgole di anni in una lista intera ordinata e unica.
def parse_years_arg(years_arg: str) -> list[int]:
    years: list[int] = []
    for token in years_arg.split(","):
        token = token.strip()
        if not token:
            continue
        years.append(int(token))
    if not years:
        raise ValueError("--years cannot be empty")
    return sorted(set(years))


def _discover_years_from_workbook(workbook_path: Path) -> list[int]:
    years: set[int] = set()
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            max_rows = min(ws.max_row or 0, 12)
            max_cols = min(ws.max_column or 0, 20)
            for row_idx in range(1, max_rows + 1):
                for col_idx in range(1, max_cols + 1):
                    year_value = _coerce_year_value(ws.cell(row=row_idx, column=col_idx).value)
                    if year_value is not None:
                        years.add(year_value)
    finally:
        wb.close()
    return sorted(years)


def resolve_years_to_process(
    args: argparse.Namespace,
    *,
    template_years: list[int],
    bdap_years: list[int],
) -> list[int]:
    """Resolve the exact years requested for the run.

    Explicit GUI/CLI selections are respected as-is. The template is used as a
    fallback only when no year selection was provided.
    """
    if args.all_years:
        return sorted(set(bdap_years))
    if args.years:
        return parse_years_arg(args.years)
    if args.year is not None:
        return [int(args.year)]
    return template_years or sorted(set(bdap_years))


def _should_use_anchor_bdap(args: argparse.Namespace) -> bool:
    """Allow --bdap as a single-year fallback, but not for multi-year discovery."""
    return args.year is not None and not args.all_years and args.years is None


def validate_inputs(args: argparse.Namespace) -> None:
    # Valida gli input richiesti in base alla modalità selezionata (anno singolo o multi-anno).
    source_analysis = args.template if getattr(args, "template", None) else args.analysis
    required_paths = [source_analysis]
    auto_multi_year_mode = args.all_years or args.years is not None or args.year is None
    if auto_multi_year_mode:
        required_paths.append(args.bdap_dir) # In multi-year mode, require the whole BDAP directory to discover files by year.
    else:
        required_paths.append(args.bdap)

    missing = [str(path) for path in required_paths if not path.exists()]
    if missing:
        missing_list = "\n".join(f"- {path}" for path in missing)
        raise FileNotFoundError(f"Missing input files:\n{missing_list}")


def apply_comune_defaults(args: argparse.Namespace, parser: argparse.ArgumentParser) -> None:
    # Applica i valori predefiniti specifici per il comune rilevati automaticamente solo per argomenti non espliciti.
    if not args.comune:
        return

    base_dir = resolve_base_dir_for_comune(args.workspace_root, args.comune)
    bdap_dir = infer_bdap_dir(base_dir)
    discovered = discover_bdap_files_by_year(bdap_dir)
    default_bdap = next(iter(discovered.values()), bdap_dir / "Rend.xlsx")

    # Mantiene inalterati i percorsi forniti dall'utente e sovrascrive solo i default del parser.
    if args.analysis == parser.get_default("analysis"):
        args.analysis = infer_new_analysis_path(base_dir)
    if args.bdap_dir == parser.get_default("bdap_dir"):
        args.bdap_dir = bdap_dir
    if args.bdap == parser.get_default("bdap"):
        args.bdap = default_bdap
    if args.output == parser.get_default("output"):
        analysis_name = Path(args.analysis)
        analysis_suffix = analysis_name.suffix or ".xlsx"
        args.output = base_dir / f"{analysis_name.stem}_compilato{analysis_suffix}"


def ensure_analysis_input(args: argparse.Namespace) -> None:
    # Verifica che l'input di analisi esista, salvo se è fornito un template esplicito.
    if getattr(args, "template", None):
        return

    if args.analysis.exists():
        return

    raise FileNotFoundError(f"Analysis file not found: {args.analysis}")


def _resolve_template_source(template_path: Path) -> Path:
    # Risolve il percorso reale del template .xlsx, gestendo l'indirezione da .gsheet esportato.
    if template_path.suffix.lower() != ".gsheet":
        return template_path

    exported_xlsx = template_path.with_suffix(".xlsx")
    if exported_xlsx.exists():
        return exported_xlsx

    raise ValueError(
        "Template .gsheet non utilizzabile direttamente. "
        "Esporta il file Google Sheets in .xlsx e riprova."
    )


def _prepare_output_from_source(args: argparse.Namespace, log: Callable[[str], None]) -> Path:
    # Copia il workbook sorgente selezionato nel percorso di output prima delle operazioni di compilazione.
    source_path = args.analysis
    if getattr(args, "template", None):
        source_path = _resolve_template_source(args.template)

    if source_path.suffix.lower() == ".gsheet":
        raise ValueError(
            "Template .gsheet non supportato. Usa un file .xlsx come template."
        )
    if not source_path.exists():
        raise FileNotFoundError(f"Missing input file: {source_path}")

    args.output.parent.mkdir(parents=True, exist_ok=True)
   
    # Copy source to output path before processing. 
    # Questo preserva il file originale e permette aggiornamenti in-place sul file di output.
    shutil.copy2(source_path, args.output) 
    
    # Se è stato usato un template specifico, registra il percorso sorgente per informare l'utente.
    if getattr(args, "template", None):
        log(f"Template used to generate output: {source_path}")
    return source_path


def run_pipeline(
    args: argparse.Namespace,
    log: Callable[[str], None] = print,
    progress: Optional[Callable[[float, int, str], None]] = None,
) -> None:
    # Esegue la pipeline end-to-end per gli anni e i fogli selezionati.
    def emit_progress(completed: float, total: int, message: str) -> None:
        if progress is None:
            return
        safe_total = max(1, int(total))
        safe_completed = min(max(0.0, float(completed)), float(safe_total))
        progress(safe_completed, safe_total, message)

    emit_progress(0, 1, "Preparazione file...")
    ensure_analysis_input(args)
    validate_inputs(args)
    _prepare_output_from_source(args, log)

    output_workbook = load_workbook(args.output, read_only=True)
    sheets_to_process = list(output_workbook.sheetnames) if getattr(args, "all_sheets", False) else [args.analysis_sheet]
    output_workbook.close()

    source_analysis = args.template if getattr(args, "template", None) else args.analysis
    template_years = _discover_years_from_workbook(source_analysis)

    bdap_by_year = discover_bdap_files_by_year(args.bdap_dir)
    years_to_process = resolve_years_to_process(
        args,
        template_years=template_years,
        bdap_years=list(bdap_by_year.keys()),
    )

    if not years_to_process:
        raise ValueError(
            "No years to process. Provide --year/--years or place BDAP .xlsx files with recognizable year in --bdap-dir (YYYY or YY->20YY for BDAP-like names)."
        )

    total_steps = max(1, len(years_to_process) * max(1, len(sheets_to_process)))
    completed_steps = 0
    emit_progress(
        completed_steps,
        total_steps,
        f"Avvio compilazione: {len(years_to_process)} anni, {len(sheets_to_process)} fogli",
    )

    # tuple (anno, celle_aggiornate, riferimenti_non_risolti) per il riepilogo finale.
    summary: list[tuple[int, int, int]] = []
    # Elabora ciascun anno in modo indipendente così un anno problematico non blocca l'intera esecuzione.
    for year in years_to_process:
        year_updated = 0
        year_unresolved = 0
        emit_progress(
            completed_steps + 0.03,
            total_steps,
            f"Anno {year}: preparazione dati BDAP",
        )
        bdap_file = bdap_by_year.get(year)
        if bdap_file is None:
            if _should_use_anchor_bdap(args) and getattr(args, "bdap", None) is not None:
                bdap_file = args.bdap
                log(
                    f"Year {year}: no year-specific BDAP file found in {args.bdap_dir}; "
                    f"using fallback anchor {bdap_file.name}."
                )
            else:
                log(
                    f"Year {year}: no BDAP file found in {args.bdap_dir}; "
                    "target cells will be left empty."
                )

        if bdap_file is not None:
            populate_fcde_sheet(args.output, bdap_file, year, log=log)

        # In modalità "tutti i fogli", ogni foglio viene trattato come un passaggio di compilazione indipendente.
        for sheet_name in sheets_to_process:
            step_start = completed_steps

            def report_row_progress(row_idx: int, max_row: int, *, year=year, sheet_name=sheet_name, step_start=step_start) -> None:
                if max_row <= 0:
                    fraction = 0.03
                else:
                    fraction = max(0.03, min(float(row_idx) / float(max_row), 1.0))
                emit_progress(
                    step_start + fraction,
                    total_steps,
                    f"Anno {year}, foglio '{sheet_name}': riga {min(row_idx, max_row)}/{max_row}",
                )

            try:
                updated, unresolved = fill_accounting_analysis(
                    analysis_path=args.output,
                    bdap_path=bdap_file,
                    output_path=args.output,
                    analysis_sheet=sheet_name,
                    year=year,
                    clear_when_missing=bdap_file is None,
                    progress=report_row_progress,
                )
                year_updated += updated
                year_unresolved += unresolved
                if getattr(args, "all_sheets", False):
                    log(
                        f"Year {year}, sheet '{sheet_name}': updated cells = {updated}; "
                        f"unresolved references = {unresolved}."
                    )
            except ValueError as exc:
                if "Year column" in str(exc) and getattr(args, "all_sheets", False):
                    log(f"Year {year}, sheet '{sheet_name}': no year column, skipped.")
                else:
                    log(f"Skipping year {year}: {exc}")
            finally:
                completed_steps += 1
                emit_progress(
                    completed_steps,
                    total_steps,
                    f"Anno {year}, foglio '{sheet_name}' completato",
                )

        summary.append((year, year_updated, year_unresolved))
        log(
            f"Year {year} populate completed: updated cells = {year_updated}; "
            f"unresolved references = {year_unresolved}."
        )

        if getattr(args, "template", None) and year_updated == 0:
            # If the year-specific pass produced no writes, retry it on a fresh
            # copy of the template and merge only that year column into the final
            # workbook. This avoids cross-year state from previous passes.
            rescue_source = source_analysis
            rescue_work = args.output.parent / f".{args.output.stem}.recheck-{year}.xlsx"
            shutil.copy2(rescue_source, rescue_work)
            rescue_bdap = bdap_by_year.get(year) or args.bdap
            fill_accounting_analysis(
                analysis_path=rescue_work,
                bdap_path=rescue_bdap,
                output_path=rescue_work,
                analysis_sheet=sheets_to_process[0],
                year=year,
                clear_when_missing=False,
            )

            master_wb = load_workbook(args.output)
            rescue_wb = load_workbook(rescue_work)
            try:
                for sheet_name in sheets_to_process:
                    master_ws = find_sheet(master_wb, sheet_name)
                    rescue_ws = find_sheet(rescue_wb, sheet_name)
                    if master_ws is None or rescue_ws is None:
                        continue
                    year_col = find_year_column(master_ws, year)
                    if year_col is None:
                        continue
                    for row_idx in range(1, master_ws.max_row + 1):
                        source_cell = rescue_ws.cell(row=row_idx, column=year_col)
                        target_cell = master_ws.cell(row=row_idx, column=year_col)
                        target_cell.value = source_cell.value
                        if source_cell.has_style:
                            target_cell._style = source_cell._style
                        if source_cell.comment is not None:
                            target_cell.comment = source_cell.comment
            finally:
                rescue_wb.close()
                master_wb.save(args.output)
                master_wb.close()
                try:
                    rescue_work.unlink()
                except Exception:
                    pass

    log(f"Final populated output: {args.output}")
    emit_progress(total_steps, total_steps, "Compilazione completata")

    if summary:
        log("\nPopulation summary by year:")
        for year, updated, unresolved in summary:
            log(f"- {year}: updated={updated}, unresolved={unresolved}")

    controlli_report = collect_controlli_post_report(args.output.parent, years=years_to_process)
    args.controlli_post_report = controlli_report
    if controlli_report.has_files:
        wrote_controlli_sheet = write_controlli_post_sheet(args.output, controlli_report)
        if not wrote_controlli_sheet:
            log("Foglio CONTROLLI-POST non trovato nel template: riepilogo non scritto nel workbook.")
        log("")
        log(format_controlli_post_summary(controlli_report))
